#!/usr/bin/env python3
"""Prepara el libro Luma e importa datos historicos trazables de prueba.

El importador nunca imprime cargas de filas ni valores personales. La
simulacion es el modo predeterminado. Las escrituras requieren --aplicar y
una URL de conexion configurada en una variable de entorno.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import os
import re
import sys
import unicodedata
import uuid
import warnings
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable
from zoneinfo import ZoneInfo

import openpyxl
import psycopg
from openpyxl.utils import get_column_letter
from psycopg.types.json import Jsonb


HEADER_ROWS = {
    "Clientes.": 3,
    "SERGIO": 3,
    "SIAM": 3,
    "lucho": 3,
    "COMPRAS": 3,
    "VENTAS": 3,
    "PAGOS F": 4,
    "INGRESOS": 4,
    "GASTOS V": 4,
    "SEGUROS": 4,
    "Resumen entradas": 1,
    "Resumen salidas": 1,
    "BUSQUEDAS": 1,
    "clientes potenciales.": 6,
    "PERSONAL": 6,
    "PROVEEDORES": 3,
}

DERIVED_SHEETS = {
    "Resumen entradas",
    "Resumen salidas",
    "BUSQUEDAS",
}

ANCHOR_COLUMNS = {
    "Clientes.": (2,),
    # Algunas filas de inventario no tienen fecha, pero contienen chasis y no
    # deben desaparecer del staging historico.
    "SERGIO": (1, 3),
    "SIAM": (1, 3),
    "lucho": (1, 8, 13),
    "COMPRAS": (1,),
    "VENTAS": (1,),
    "PAGOS F": (2,),
    "INGRESOS": (2,),
    "GASTOS V": (2,),
    "SEGUROS": (3,),
    "clientes potenciales.": (2, 3, 9, 10, 16, 17),
    "PERSONAL": (1,),
    "PROVEEDORES": (2,),
}

DEFERRED_SHEETS = {
    "lucho": "MAPEO_MULTIBLOQUE_REQUERIDO",
    "PAGOS F": "PAGO_CONCILIACION_REQUERIDA",
    "INGRESOS": "COBRANZA_CONCILIACION_REQUERIDA",
    "GASTOS V": "GASTO_CONCILIACION_REQUERIDA",
    "SEGUROS": "MODULO_SEGUROS_DIFERIDO",
    "clientes potenciales.": "MODULO_PROSPECTOS_DIFERIDO",
}

ARGENTINA_TZ = ZoneInfo("America/Argentina/Buenos_Aires")
IMPORT_NAMESPACE = uuid.UUID("026672d7-a54a-44e4-929f-03713b7d782a")
TAMANO_LOTE = 250
IMPORT_DATABASE_ENV_VAR = "DIRECT_URL"
LEGACY_IMPORT_MIGRATION = "20260829130000_legacy_import_support"
SENTINEL_VINS = {
    "USADA",
    "USADO",
    "SENA",
    "RESERVA",
    "SINVIN",
    "SINDATO",
    "PENDIENTE",
    "NA",
}


@dataclass(frozen=True)
class SourceRef:
    sheet: str
    row: int
    block: str = "predeterminado"


@dataclass
class StagedRow:
    source: SourceRef
    payload: dict[str, Any]
    hash_original: str
    fila_importacion_id: uuid.UUID | None = None
    estado: str = "PENDIENTE"
    errors: list[str] = field(default_factory=list)
    targets: list[dict[str, str]] = field(default_factory=list)


@dataclass
class CustomerCandidate:
    source: SourceRef
    tipo_documento: str
    numero_documento: str
    documento_normalizado: str
    nombre_completo: str
    nombre_normalizado: str
    telefono: str | None


@dataclass
class StaffCandidate:
    source: SourceRef
    nombre_completo: str
    nombre_normalizado: str
    telefono: str | None
    direccion: str | None
    role_name: str | None


@dataclass
class SupplierCandidate:
    source: SourceRef
    razon_social: str
    nombre_normalizado: str
    direccion: str | None
    nombre_contacto: str | None
    telefono: str | None
    notas: str | None


@dataclass
class VehicleCandidate:
    source: SourceRef
    tipo_vehiculo: str
    brand_name: str
    brand_normalized: str
    model_name: str
    model_normalized: str
    vin_mostrado: str | None
    vin_normalizado: str | None
    numero_motor: str | None
    engine_normalized: str | None
    anio_fabricacion: int | None
    color: str | None
    supplier_name: str | None
    supplier_normalized: str | None
    recibido_en: datetime | None
    costo_compra: Decimal | None
    estado_inventario: str
    errors: list[str] = field(default_factory=list)


@dataclass
class GastoHistorico:
    source: StagedRow
    categoria: str
    detalle: str
    fecha_generacion: date
    importe: Decimal
    recuperable: bool
    estado_pago: str
    pagador_original: str | None
    recuperable_original: str | None
    referencia_origen: str | None
    vin_origen_mostrado: str | None
    vin_origen_normalizado: str | None
    unidad_vehiculo_id: uuid.UUID | None
    datos_inferidos: dict[str, Any]
    cuenta_caja_id: uuid.UUID | None


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Lee el XLSX de Luma, prepara filas de origen e importa datos "
            "maestros deterministas. --datos-prueba agrega el historico "
            "trazable despues de aplicar la migracion Prisma de importacion."
        )
    )
    parser.add_argument("--libro", required=True, type=Path)
    parser.add_argument(
        "--organizacion",
        default=None,
        help="Codigo de la organizacion activa que recibira la importacion.",
    )
    parser.add_argument(
        "--variable-entorno-base-datos",
        default=IMPORT_DATABASE_ENV_VAR,
        help="Variable de entorno que contiene la URL directa de PostgreSQL.",
    )
    parser.add_argument(
        "--sucursal-predeterminada",
        help=(
            "Sucursal asignada a las unidades de COMPRAS porque el libro no tiene columna de sucursal. Las unidades no se importan si se omite."
        ),
    )
    parser.add_argument(
        "--aplicar",
        action="store_true",
        help=(
            "Escribe staging y filas maestras seguras. Con --datos-prueba "
            "tambien escribe el historico trazable. Sin esta opcion solo se simula."
        ),
    )
    parser.add_argument(
        "--solo-staging",
        action="store_true",
        help="Con --aplicar, escribe lotes y filas de importacion sin filas de negocio.",
    )
    parser.add_argument(
        "--datos-prueba",
        action="store_true",
        help=(
            "Con --aplicar, importa el historico del libro como datos de "
            "prueba trazables. Requiere --organizacion y "
            "--sucursal-predeterminada."
        ),
    )
    return parser.parse_args()


def safe_text(value: Any) -> str | None:
    if value is None:
        return None
    text = re.sub(r"\s+", " ", str(value)).strip()
    return text or None


def normalize_text(value: Any) -> str:
    text = safe_text(value) or ""
    decomposed = unicodedata.normalize("NFD", text)
    ascii_text = "".join(
        char for char in decomposed if unicodedata.category(char) != "Mn"
    )
    return re.sub(r"\s+", " ", ascii_text).strip().lower()


def normalize_identifier(value: Any) -> str:
    text = normalize_text(value).upper()
    return re.sub(r"[^A-Z0-9]", "", text)


def normalize_document(value: Any) -> str | None:
    text = safe_text(value)
    if text is None:
        return None
    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
    digits = re.sub(r"\D", "", text)
    return digits if 6 <= len(digits) <= 11 else None


def tipo_documento(document: str) -> str:
    return "CUIT" if len(document) == 11 else "DNI"


def valid_vin(value: Any) -> tuple[str | None, str | None]:
    display = safe_text(value)
    normalized = normalize_identifier(value)
    if not display or not normalized:
        return None, "VIN_FALTANTE"
    if normalized in SENTINEL_VINS or "USADA" in normalized or "SENA" in normalized:
        return None, "VIN_MARCADOR"
    if not 6 <= len(normalized) <= 32:
        return None, "VIN_LONGITUD_INVALIDA"
    return normalized, None


def parse_year(value: Any) -> tuple[int | None, str | None]:
    if value in (None, ""):
        return None, None
    try:
        year = int(float(value))
    except (TypeError, ValueError):
        return None, "ANIO_INVALIDO"
    if not 1886 <= year <= 2100:
        return None, "ANIO_INVALIDO"
    return year, None


def parse_amount(value: Any) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        if isinstance(value, Decimal):
            amount = value
        elif isinstance(value, (int, float)):
            amount = Decimal(str(value))
        else:
            text = safe_text(value)
            if text is None:
                return None
            normalized = (
                text.replace("$", "")
                .replace("ARS", "")
                .replace("\u00a0", "")
                .replace(" ", "")
            )
            if "," in normalized and "." in normalized:
                if normalized.rfind(",") > normalized.rfind("."):
                    normalized = normalized.replace(".", "").replace(",", ".")
                else:
                    normalized = normalized.replace(",", "")
            elif "," in normalized:
                whole, fraction = normalized.rsplit(",", 1)
                normalized = (
                    f"{whole}.{fraction}"
                    if len(fraction) in (1, 2)
                    else normalized.replace(",", "")
                )
            amount = Decimal(normalized)
    except (InvalidOperation, TypeError, ValueError):
        return None
    if not amount.is_finite() or amount < 0:
        return None
    return amount.quantize(Decimal("0.01"))


def parse_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        if value.tzinfo is None:
            return value.replace(tzinfo=ARGENTINA_TZ)
        return value
    if isinstance(value, date):
        return datetime.combine(value, time.min, tzinfo=ARGENTINA_TZ)
    text = safe_text(value)
    if text:
        for formato in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(text, formato).replace(tzinfo=ARGENTINA_TZ)
            except ValueError:
                continue
    return None


def parse_date(value: Any) -> date | None:
    parsed = parse_datetime(value)
    return parsed.date() if parsed else None


def positive_amount(value: Any) -> Decimal | None:
    amount = parse_amount(value)
    return amount if amount is not None and amount > 0 else None


def append_error(staged_row: StagedRow, code: str) -> None:
    if code not in staged_row.errors:
        staged_row.errors.append(code)


def limited_text(value: Any, maximum: int) -> str | None:
    text = safe_text(value)
    return text[:maximum] if text else None


def inferred_data(**values: Any) -> dict[str, Any]:
    return {
        key: json_value(value)
        for key, value in values.items()
        if value not in (None, "")
    }


def estado_pago_desde_origen(value: Any) -> str:
    normalized = normalize_text(value)
    if normalized in {"pagado", "cobrado", "cancelado"}:
        return "PAGADO" if normalized != "cancelado" else "CANCELADA"
    if normalized in {"pago parcial", "parcial"}:
        return "PAGO_PARCIAL"
    if normalized == "vencido":
        return "VENCIDO"
    if normalized in {"reintegrado", "devuelto"}:
        return "REINTEGRADO"
    return "PENDIENTE"


def es_estado_cobrado(value: Any) -> bool:
    return normalize_text(value) in {"cobrado", "pagado"}


def es_transferencia_origen(value: Any) -> bool:
    return "transferencia" in normalize_text(value)


def booleano_desde_origen(value: Any) -> bool | None:
    if isinstance(value, bool):
        return value
    normalized = normalize_text(value)
    if normalized in {"si", "sí", "s", "x", "true", "verdadero", "recuperada"}:
        return True
    if normalized in {"no", "n", "false", "falso"}:
        return False
    return None


def fecha_historica(fecha: date) -> datetime:
    return datetime.combine(fecha, time.min, tzinfo=ARGENTINA_TZ)


def json_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, dict):
        return {str(key): json_value(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [json_value(item) for item in value]
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return None
        return int(value) if value.is_integer() else value
    if isinstance(value, (str, int, bool)):
        return value
    return str(value)


def canonical_json(payload: dict[str, Any]) -> str:
    return json.dumps(
        payload,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def row_payload(
    value_sheet: openpyxl.worksheet.worksheet.Worksheet,
    formula_sheet: openpyxl.worksheet.worksheet.Worksheet,
    row_number: int,
    header_row: int,
) -> tuple[dict[str, Any], bool]:
    payload: dict[str, Any] = {}
    has_direct_value = False
    seen_headers: Counter[str] = Counter()

    for column in range(1, value_sheet.max_column + 1):
        formula_cell = formula_sheet.cell(row=row_number, column=column)
        value = value_sheet.cell(row=row_number, column=column).value
        formula_value = formula_cell.value
        if value in (None, "") and formula_value in (None, ""):
            continue

        header = safe_text(value_sheet.cell(row=header_row, column=column).value)
        if not header:
            header = f"column_{get_column_letter(column)}"
        seen_headers[header] += 1
        key = header if seen_headers[header] == 1 else f"{header}_{seen_headers[header]}"
        payload[key] = json_value(value)
        has_direct_value = has_direct_value or formula_cell.data_type != "f"

    return payload, has_direct_value


def load_staged_rows(
    workbook_values: openpyxl.Workbook,
    workbook_formulas: openpyxl.Workbook,
) -> list[StagedRow]:
    staged: list[StagedRow] = []
    for nombre_hoja in workbook_values.sheetnames:
        if nombre_hoja not in HEADER_ROWS:
            raise ValueError(f"Hoja de libro no admitida: {nombre_hoja}")
        if nombre_hoja in DERIVED_SHEETS:
            continue

        values = workbook_values[nombre_hoja]
        formulas = workbook_formulas[nombre_hoja]
        header_row = HEADER_ROWS[nombre_hoja]
        for row_number in range(header_row + 1, values.max_row + 1):
            anchored = any(
                values.cell(row=row_number, column=column).value not in (None, "")
                and formulas.cell(row=row_number, column=column).data_type != "f"
                for column in ANCHOR_COLUMNS[nombre_hoja]
            )
            if not anchored:
                continue
            payload, has_direct_value = row_payload(
                values,
                formulas,
                row_number,
                header_row,
            )
            if not payload or not has_direct_value:
                continue

            raw_json = canonical_json(payload)
            estado = "OMITIDA" if nombre_hoja in DERIVED_SHEETS else "PENDIENTE"
            staged.append(
                StagedRow(
                    source=SourceRef(nombre_hoja, row_number),
                    payload=payload,
                    hash_original=hashlib.sha256(raw_json.encode("utf-8")).hexdigest(),
                    estado=estado,
                )
            )
    return staged


def cell(
    workbook: openpyxl.Workbook,
    sheet: str,
    row: int,
    column: int,
) -> Any:
    return workbook[sheet].cell(row=row, column=column).value


def iter_source_rows(
    staged: Iterable[StagedRow],
    sheet: str,
) -> Iterable[StagedRow]:
    return (row for row in staged if row.source.sheet == sheet)


def extract_customers(
    workbook: openpyxl.Workbook,
    staged: list[StagedRow],
) -> list[CustomerCandidate]:
    candidates: list[CustomerCandidate] = []
    layouts = {
        "Clientes.": (2, 3, 8),
        "VENTAS": (6, 7, None),
        "SEGUROS": (1, 2, None),
    }
    for sheet, (document_column, name_column, phone_column) in layouts.items():
        for staged_row in iter_source_rows(staged, sheet):
            row = staged_row.source.row
            document = normalize_document(cell(workbook, sheet, row, document_column))
            nombre = safe_text(cell(workbook, sheet, row, name_column))
            telefono = (
                safe_text(cell(workbook, sheet, row, phone_column))
                if phone_column
                else None
            )
            if not document:
                staged_row.errors.append("CLIENTE_DOCUMENTO_INVALIDO")
                continue
            if not nombre:
                staged_row.errors.append("CLIENTE_NOMBRE_FALTANTE")
                continue
            candidates.append(
                CustomerCandidate(
                    source=staged_row.source,
                    tipo_documento=tipo_documento(document),
                    numero_documento=document,
                    documento_normalizado=document,
                    nombre_completo=nombre,
                    nombre_normalizado=normalize_text(nombre),
                    telefono=telefono,
                )
            )
    return candidates


def extract_staff(
    workbook: openpyxl.Workbook,
    staged: list[StagedRow],
) -> list[StaffCandidate]:
    candidates: list[StaffCandidate] = []
    for staged_row in iter_source_rows(staged, "PERSONAL"):
        row = staged_row.source.row
        nombre = safe_text(cell(workbook, "PERSONAL", row, 1))
        if not nombre:
            staged_row.errors.append("PERSONAL_NOMBRE_FALTANTE")
            continue
        candidates.append(
            StaffCandidate(
                source=staged_row.source,
                nombre_completo=nombre,
                nombre_normalizado=normalize_text(nombre),
                direccion=safe_text(cell(workbook, "PERSONAL", row, 2)),
                telefono=safe_text(cell(workbook, "PERSONAL", row, 3)),
                role_name=safe_text(cell(workbook, "PERSONAL", row, 4)),
            )
        )
    return candidates


def flag_customer_identity_conflicts(
    candidates: list[CustomerCandidate],
    staged_by_source: dict[SourceRef, StagedRow],
) -> None:
    grouped: dict[tuple[str, str], list[CustomerCandidate]] = defaultdict(list)
    for candidate in candidates:
        grouped[(candidate.tipo_documento, candidate.documento_normalizado)].append(
            candidate
        )
    for group in grouped.values():
        if len({candidate.nombre_normalizado for candidate in group}) <= 1:
            continue
        for candidate in group:
            errors = staged_by_source[candidate.source].errors
            if "CLIENTE_IDENTIDAD_CONFLICTO" not in errors:
                errors.append("CLIENTE_IDENTIDAD_CONFLICTO")


def extract_suppliers(
    workbook: openpyxl.Workbook,
    staged: list[StagedRow],
) -> list[SupplierCandidate]:
    candidates: dict[str, SupplierCandidate] = {}

    for staged_row in iter_source_rows(staged, "PROVEEDORES"):
        row = staged_row.source.row
        nombre = safe_text(cell(workbook, "PROVEEDORES", row, 2))
        if not nombre:
            continue
        normalized = normalize_text(nombre)
        candidates.setdefault(
            normalized,
            SupplierCandidate(
                source=staged_row.source,
                razon_social=nombre,
                nombre_normalizado=normalized,
                direccion=safe_text(cell(workbook, "PROVEEDORES", row, 3)),
                nombre_contacto=safe_text(cell(workbook, "PROVEEDORES", row, 4)),
                telefono=safe_text(cell(workbook, "PROVEEDORES", row, 5)),
                notas=safe_text(cell(workbook, "PROVEEDORES", row, 8)),
            ),
        )

    for staged_row in iter_source_rows(staged, "COMPRAS"):
        row = staged_row.source.row
        nombre = safe_text(cell(workbook, "COMPRAS", row, 10))
        if not nombre:
            continue
        normalized = normalize_text(nombre)
        candidates.setdefault(
            normalized,
            SupplierCandidate(
                source=staged_row.source,
                razon_social=nombre,
                nombre_normalizado=normalized,
                direccion=None,
                nombre_contacto=None,
                telefono=None,
                notas=None,
            ),
        )

    return list(candidates.values())


def inventory_status_by_vin(
    workbook: openpyxl.Workbook,
    staged: list[StagedRow],
) -> dict[str, str]:
    statuses: dict[str, set[str]] = defaultdict(set)

    for sheet in ("SERGIO", "SIAM"):
        for staged_row in iter_source_rows(staged, sheet):
            row = staged_row.source.row
            vin, error = valid_vin(cell(workbook, sheet, row, 3))
            if error:
                continue
            estado = normalize_text(cell(workbook, sheet, row, 5))
            if estado == "vendida":
                statuses[vin].add("VENDIDO")
            elif estado == "stock":
                statuses[vin].add("EN_STOCK")

    for staged_row in iter_source_rows(staged, "VENTAS"):
        row = staged_row.source.row
        vin, error = valid_vin(cell(workbook, "VENTAS", row, 2))
        if not error:
            statuses[vin].add("VENDIDO")

    resolved: dict[str, str] = {}
    for vin, values in statuses.items():
        resolved[vin] = "VENDIDO" if "VENDIDO" in values else next(iter(values))
    return resolved


def tipo_vehiculo_ventas_por_vin(
    workbook: openpyxl.Workbook,
    staged: list[StagedRow],
) -> dict[str, str]:
    tipos: dict[str, set[str]] = defaultdict(set)
    for staged_row in iter_source_rows(staged, "VENTAS"):
        row = staged_row.source.row
        vin, error = valid_vin(cell(workbook, "VENTAS", row, 2))
        if error or vin is None:
            continue
        tipo = normalize_text(cell(workbook, "VENTAS", row, 3))
        if "moto" in tipo:
            tipos[vin].add("MOTO")
        elif "auto" in tipo:
            tipos[vin].add("AUTO")
    return {
        vin: next(iter(valores))
        for vin, valores in tipos.items()
        if len(valores) == 1
    }


def extract_vehicles(
    workbook: openpyxl.Workbook,
    staged: list[StagedRow],
) -> list[VehicleCandidate]:
    candidates: list[VehicleCandidate] = []
    inventory_statuses = inventory_status_by_vin(workbook, staged)
    tipos_ventas = tipo_vehiculo_ventas_por_vin(workbook, staged)

    for staged_row in iter_source_rows(staged, "COMPRAS"):
        row = staged_row.source.row
        raw_type = normalize_text(cell(workbook, "COMPRAS", row, 2))
        brand = safe_text(cell(workbook, "COMPRAS", row, 4))
        model = safe_text(cell(workbook, "COMPRAS", row, 5))
        vin_column_3, vin_error_3 = valid_vin(cell(workbook, "COMPRAS", row, 3))
        vin_column_8, vin_error_8 = valid_vin(cell(workbook, "COMPRAS", row, 8))
        errors: list[str] = []
        if "moto" in raw_type:
            tipo_vehiculo = "MOTO"
        elif "auto" in raw_type:
            tipo_vehiculo = "AUTO"
        else:
            vin_para_tipo = (
                vin_column_3
                if vin_column_3 == vin_column_8
                else vin_column_3 or vin_column_8
            )
            tipo_vehiculo = tipos_ventas.get(vin_para_tipo or "")
            if tipo_vehiculo is None:
                append_error(staged_row, "VEHICULO_TIPO_DESCONOCIDO")
                continue
            errors.append("VEHICULO_TIPO_INFERIDO_DE_VENTAS")
        if not brand or not model:
            append_error(staged_row, "CATALOGO_DATOS_FALTANTES")
            continue

        vin: str | None
        vin_mostrado: str | None
        if vin_column_3 and vin_column_8 and vin_column_3 != vin_column_8:
            vin = None
            vin_mostrado = None
            errors.append("VIN_COLUMNAS_CONFLICTO")
        else:
            vin = vin_column_3 or vin_column_8
            source_column = 3 if vin_column_3 else 8
            vin_mostrado = safe_text(cell(workbook, "COMPRAS", row, source_column))
            if not vin:
                errors.append(vin_error_3 or vin_error_8 or "VIN_FALTANTE")

        year, year_error = parse_year(cell(workbook, "COMPRAS", row, 7))
        if year_error:
            errors.append(year_error)

        engine = safe_text(cell(workbook, "COMPRAS", row, 9))
        engine_normalized = normalize_identifier(engine) or None
        if engine and not engine_normalized:
            errors.append("MOTOR_NUMERO_INVALIDO")
        recibido_en = parse_datetime(cell(workbook, "COMPRAS", row, 1))
        if recibido_en is None:
            errors.append("FECHA_COMPRA_INVALIDA")

        supplier_name = safe_text(cell(workbook, "COMPRAS", row, 10))
        costo_compra = parse_amount(cell(workbook, "COMPRAS", row, 17))
        estado_inventario = inventory_statuses.get(vin or "", "BLOQUEADO")
        if estado_inventario == "BLOQUEADO":
            errors.append("ESTADO_INVENTARIO_SIN_CONFIRMAR")
        staged_row.errors.extend(
            error for error in errors if error not in staged_row.errors
        )

        candidates.append(
            VehicleCandidate(
                source=staged_row.source,
                tipo_vehiculo=tipo_vehiculo,
                brand_name=brand,
                brand_normalized=normalize_text(brand),
                model_name=model,
                model_normalized=normalize_text(model),
                vin_mostrado=vin_mostrado,
                vin_normalizado=vin,
                numero_motor=engine,
                engine_normalized=engine_normalized,
                anio_fabricacion=year,
                color=safe_text(cell(workbook, "COMPRAS", row, 6)),
                supplier_name=supplier_name,
                supplier_normalized=(
                    normalize_text(supplier_name) if supplier_name else None
                ),
                recibido_en=recibido_en,
                costo_compra=costo_compra,
                estado_inventario=estado_inventario,
                errors=errors,
            )
        )
    return candidates


def deterministic_id(kind: str, key: str) -> uuid.UUID:
    return uuid.uuid5(IMPORT_NAMESPACE, f"{kind}:{key}")


def mark_target(
    staged_by_source: dict[SourceRef, StagedRow],
    source: SourceRef,
    table: str,
    target_id: uuid.UUID,
    bloque: str | None = None,
) -> None:
    target = {"tabla": table, "id": str(target_id)}
    if bloque:
        target["bloque"] = bloque
    if target not in staged_by_source[source].targets:
        staged_by_source[source].targets.append(target)


def ensure_import_schema(
    connection: psycopg.Connection[Any], datos_prueba: bool = False
) -> None:
    required = [
        "organizaciones",
        "catalogo_organizaciones",
        "lotes_importacion",
        "filas_importacion",
        "clientes",
        "personal",
        "proveedores",
        "marcas_vehiculos",
        "modelos_vehiculos",
        "versiones_vehiculos",
        "unidades_vehiculos",
    ]
    if datos_prueba:
        required.extend(
            (
                "operaciones",
                "componentes_pago_operacion",
                "cuentas_caja",
                "movimientos_caja",
                "gastos",
                "ingresos",
                "polizas_seguros",
                "prospectos",
                "registros_inventario_importados",
            )
        )
    with connection.cursor() as cursor:
        for table in required:
            cursor.execute("SELECT to_regclass(%s)", (f"public.{table}",))
            if cursor.fetchone()[0] is None:
                raise RuntimeError(
                    f"Falta la tabla requerida public.{table}. "
                    "No ejecute database/001..003 manualmente. Verifique "
                    "`npx prisma migrate status` y aplique las migraciones "
                    "versionadas con `npx prisma migrate deploy` solo despues "
                    "de reconciliar el historial de la base."
                )
        if datos_prueba:
            columnas_requeridas = (
                ("personal", "es_actor_sistema_importado"),
                ("clientes", "fila_importacion_id"),
                ("cuentas_caja", "es_importada"),
                ("unidades_vehiculos", "fila_importacion_id"),
                ("operaciones", "precios_referencia_completos"),
                ("componentes_pago_operacion", "fila_importacion_id"),
                ("gastos", "vin_origen_normalizado"),
                ("movimientos_caja", "ingreso_id"),
            )
            for tabla, columna in columnas_requeridas:
                cursor.execute(
                    """
                    SELECT 1
                    FROM information_schema.columns
                    WHERE table_schema = 'public'
                      AND table_name = %s
                      AND column_name = %s
                    """,
                    (tabla, columna),
                )
                if cursor.fetchone() is None:
                    raise RuntimeError(
                        "Faltan columnas de importacion legacy. Verifique que "
                        f"la migracion Prisma {LEGACY_IMPORT_MIGRATION} figure "
                        "como aplicada; no ejecute database/003 manualmente."
                    )


def resolve_organizacion(
    connection: psycopg.Connection[Any], codigo: str
) -> tuple[uuid.UUID, str]:
    with connection.cursor() as cursor:
        cursor.execute(
            """
            SELECT id, codigo, tipo
            FROM organizaciones
            WHERE codigo = %s AND activa
            """,
            (codigo,),
        )
        row = cursor.fetchone()
        if row is None:
            raise RuntimeError(f"Organizacion activa no encontrada: {codigo}")
        organizacion_id, codigo_resuelto, tipo_organizacion = row
        acceso_global = "true" if tipo_organizacion == "CASA_CENTRAL" else "false"
        cursor.execute(
            """
            SELECT set_config('app.organizacion_id', %s, true),
                   set_config('app.acceso_global', %s, true)
            """,
            (str(organizacion_id), acceso_global),
        )
    return organizacion_id, codigo_resuelto


def stage_rows(
    connection: psycopg.Connection[Any],
    workbook: Path,
    workbook_hash: str,
    staged: list[StagedRow],
    organizacion_id: uuid.UUID,
) -> uuid.UUID:
    with connection.cursor() as cursor:
        cursor.execute(
            """
            INSERT INTO lotes_importacion (
              organizacion_id,
              nombre_archivo_origen,
              sha256_origen,
              estado,
              total_filas,
              iniciado_en
            )
            VALUES (%s, %s, %s, 'PROCESANDO', %s, now())
            ON CONFLICT (organizacion_id, sha256_origen) DO UPDATE
            SET nombre_archivo_origen = EXCLUDED.nombre_archivo_origen,
                estado = 'PROCESANDO',
                total_filas = EXCLUDED.total_filas,
                iniciado_en = now(),
                finalizado_en = NULL,
                actualizado_en = now()
            RETURNING id
            """,
            (organizacion_id, workbook.name, workbook_hash, len(staged)),
        )
        lote_id = cursor.fetchone()[0]

        ejecutar_lote(
            cursor,
            """
            INSERT INTO filas_importacion (
              organizacion_id,
              lote_id,
              nombre_hoja,
              nombre_bloque,
              fila_origen,
              carga_original,
              hash_original,
              estado,
              codigos_error,
              referencias_destino
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (lote_id, nombre_hoja, nombre_bloque, fila_origen)
            DO UPDATE SET
              carga_original = EXCLUDED.carga_original,
              hash_original = EXCLUDED.hash_original,
              estado = EXCLUDED.estado,
              codigos_error = EXCLUDED.codigos_error,
              referencias_destino = EXCLUDED.referencias_destino,
              actualizado_en = now()
            """,
            [
                (
                    organizacion_id,
                    lote_id,
                    row.source.sheet,
                    row.source.block,
                    row.source.row,
                    Jsonb(row.payload),
                    row.hash_original,
                    row.estado,
                    Jsonb(row.errors),
                    Jsonb(row.targets),
                )
                for row in staged
            ],
        )
        cursor.execute(
            """
            SELECT id, nombre_hoja, nombre_bloque, fila_origen
            FROM filas_importacion
            WHERE lote_id = %s
            """,
            (lote_id,),
        )
        ids_por_origen = {
            SourceRef(nombre_hoja, fila_origen, nombre_bloque): fila_id
            for fila_id, nombre_hoja, nombre_bloque, fila_origen in cursor.fetchall()
        }
        for row in staged:
            row.fila_importacion_id = ids_por_origen.get(row.source)
            if row.fila_importacion_id is None:
                raise RuntimeError(
                    "No se pudo resolver la fila de staging para una fila del libro."
                )
    return lote_id


def import_staff(
    connection: psycopg.Connection[Any],
    candidates: list[StaffCandidate],
    staged_by_source: dict[SourceRef, StagedRow],
    organizacion_id: uuid.UUID,
    codigo_organizacion: str,
    sucursal_principal_id: uuid.UUID | None = None,
) -> int:
    imported = 0
    with connection.cursor() as cursor:
        cursor.execute("SELECT id, lower(trim(nombre)) FROM roles")
        role_ids = {nombre: rol_id for rol_id, nombre in cursor.fetchall()}

        for candidate in candidates:
            target_id = deterministic_id(
                "staff", f"{organizacion_id}:{codigo_organizacion}:{candidate.nombre_normalizado}"
            )
            rol_id = role_ids.get(normalize_text(candidate.role_name))
            cursor.execute(
                """
                INSERT INTO personal (
                  id,
                  organizacion_id,
                  nombre_completo,
                  nombre_normalizado,
                  telefono,
                  direccion,
                  sucursal_principal_id,
                  rol_id,
                  puede_iniciar_sesion,
                  estado
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, false, 'ACTIVO')
                ON CONFLICT (id) DO UPDATE SET
                  nombre_completo = EXCLUDED.nombre_completo,
                  telefono = coalesce(EXCLUDED.telefono, personal.telefono),
                  direccion = coalesce(EXCLUDED.direccion, personal.direccion),
                  sucursal_principal_id = coalesce(
                    EXCLUDED.sucursal_principal_id,
                    personal.sucursal_principal_id
                  ),
                  rol_id = coalesce(EXCLUDED.rol_id, personal.rol_id),
                  actualizado_en = now()
                """,
                (
                    target_id,
                    organizacion_id,
                    candidate.nombre_completo,
                    candidate.nombre_normalizado,
                    candidate.telefono,
                    candidate.direccion,
                    sucursal_principal_id,
                    rol_id,
                ),
            )
            mark_target(staged_by_source, candidate.source, "personal", target_id)
            imported += 1
    return imported


def import_suppliers(
    connection: psycopg.Connection[Any],
    candidates: list[SupplierCandidate],
    staged_by_source: dict[SourceRef, StagedRow],
    organizacion_id: uuid.UUID,
    codigo_organizacion: str,
) -> dict[str, uuid.UUID]:
    supplier_ids: dict[str, uuid.UUID] = {}
    with connection.cursor() as cursor:
        for candidate in candidates:
            target_id = deterministic_id(
                "supplier", f"{organizacion_id}:{codigo_organizacion}:{candidate.nombre_normalizado}"
            )
            cursor.execute(
                """
                INSERT INTO proveedores (
                  id,
                  organizacion_id,
                  razon_social,
                  nombre_normalizado,
                  direccion,
                  nombre_contacto,
                  telefono,
                  notas,
                  activo
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, true)
                ON CONFLICT (organizacion_id, nombre_normalizado) DO UPDATE SET
                  direccion = coalesce(EXCLUDED.direccion, proveedores.direccion),
                  nombre_contacto = coalesce(EXCLUDED.nombre_contacto, proveedores.nombre_contacto),
                  telefono = coalesce(EXCLUDED.telefono, proveedores.telefono),
                  notas = coalesce(EXCLUDED.notas, proveedores.notas),
                  actualizado_en = now()
                RETURNING id
                """,
                (
                    target_id,
                    organizacion_id,
                    candidate.razon_social,
                    candidate.nombre_normalizado,
                    candidate.direccion,
                    candidate.nombre_contacto,
                    candidate.telefono,
                    candidate.notas,
                ),
            )
            actual_id = cursor.fetchone()[0]
            supplier_ids[candidate.nombre_normalizado] = actual_id
            mark_target(staged_by_source, candidate.source, "proveedores", actual_id)
    return supplier_ids


def import_customers(
    connection: psycopg.Connection[Any],
    candidates: list[CustomerCandidate],
    staged_by_source: dict[SourceRef, StagedRow],
    organizacion_id: uuid.UUID,
    codigo_organizacion: str,
) -> int:
    grouped: dict[tuple[str, str], list[CustomerCandidate]] = defaultdict(list)
    for candidate in candidates:
        grouped[(candidate.tipo_documento, candidate.documento_normalizado)].append(
            candidate
        )

    imported = 0
    with connection.cursor() as cursor:
        for key, group in grouped.items():
            names = {candidate.nombre_normalizado for candidate in group}
            if len(names) > 1:
                for candidate in group:
                    staged_by_source[candidate.source].errors.append(
                        "CLIENTE_IDENTIDAD_CONFLICTO"
                    )
                continue

            candidate = group[0]
            target_id = deterministic_id(
                "customer",
                f"{organizacion_id}:{codigo_organizacion}:{key[0]}:{key[1]}",
            )
            telefono = next((item.telefono for item in group if item.telefono), None)
            cursor.execute(
                """
                INSERT INTO clientes (
                  id,
                  organizacion_id,
                  tipo_documento,
                  numero_documento,
                  documento_normalizado,
                  nombre_completo,
                  nombre_normalizado,
                  telefono,
                  activo
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, true)
                ON CONFLICT (organizacion_id, tipo_documento, documento_normalizado)
                  WHERE documento_normalizado IS NOT NULL
                DO UPDATE SET
                  telefono = coalesce(clientes.telefono, EXCLUDED.telefono),
                  actualizado_en = now()
                RETURNING id
                """,
                (
                    target_id,
                    organizacion_id,
                    candidate.tipo_documento,
                    candidate.numero_documento,
                    candidate.documento_normalizado,
                    candidate.nombre_completo,
                    candidate.nombre_normalizado,
                    telefono,
                ),
            )
            actual_id = cursor.fetchone()[0]
            for item in group:
                mark_target(staged_by_source, item.source, "clientes", actual_id)
            imported += 1
    return imported


def import_catalog_and_units(
    connection: psycopg.Connection[Any],
    candidates: list[VehicleCandidate],
    staged_by_source: dict[SourceRef, StagedRow],
    supplier_ids: dict[str, uuid.UUID],
    default_branch: str | None,
    organizacion_id: uuid.UUID,
    codigo_organizacion: str,
    datos_prueba: bool = False,
) -> tuple[int, int, int]:
    catalog_count = 0
    unit_count = 0
    quarantined_units = 0
    sucursal_id: uuid.UUID | None = None

    with connection.cursor() as cursor:
        if default_branch:
            cursor.execute(
                """
                SELECT id FROM sucursales
                WHERE organizacion_id = %s
                  AND lower(trim(nombre)) = lower(trim(%s))
                """,
                (organizacion_id, default_branch),
            )
            row = cursor.fetchone()
            if row is None:
                raise RuntimeError(f"Sucursal no encontrada: {default_branch}")
            sucursal_id = row[0]

        catalog_cache: dict[tuple[str, str, str], uuid.UUID] = {}
        es_casa_central = codigo_organizacion == "LUMA_CENTRAL"
        for candidate in candidates:
            catalog_key = (
                candidate.tipo_vehiculo,
                candidate.brand_normalized,
                candidate.model_normalized,
            )
            version_id = catalog_cache.get(catalog_key)
            if version_id is None:
                marca_id = deterministic_id("brand", candidate.brand_normalized)
                if es_casa_central:
                    cursor.execute(
                        """
                        INSERT INTO marcas_vehiculos (id, nombre, nombre_normalizado)
                        VALUES (%s, %s, %s)
                        ON CONFLICT (nombre_normalizado) DO NOTHING
                        RETURNING id
                        """,
                        (
                            marca_id,
                            candidate.brand_name,
                            candidate.brand_normalized,
                        ),
                    )
                    marca_insertada = cursor.fetchone()
                else:
                    marca_insertada = None
                if marca_insertada is not None:
                    marca_id = marca_insertada[0]
                else:
                    cursor.execute(
                        """
                        SELECT id
                        FROM marcas_vehiculos
                        WHERE nombre_normalizado = %s
                        """,
                        (candidate.brand_normalized,),
                    )
                    marca_existente = cursor.fetchone()
                    if marca_existente is None:
                        append_error(
                            staged_by_source[candidate.source],
                            "CATALOGO_MARCA_REQUIERE_CASA_CENTRAL",
                        )
                        quarantined_units += 1
                        continue
                    marca_id = marca_existente[0]
                modelo_id = deterministic_id(
                    "model",
                    ":".join(
                        (
                            str(marca_id),
                            candidate.tipo_vehiculo,
                            candidate.model_normalized,
                        )
                    ),
                )
                if es_casa_central:
                    cursor.execute(
                        """
                        INSERT INTO modelos_vehiculos (
                          id,
                          marca_id,
                          tipo_vehiculo,
                          nombre,
                          nombre_normalizado
                        )
                        VALUES (%s, %s, %s, %s, %s)
                        ON CONFLICT (marca_id, tipo_vehiculo, nombre_normalizado)
                        DO NOTHING
                        RETURNING id
                        """,
                        (
                            modelo_id,
                            marca_id,
                            candidate.tipo_vehiculo,
                            candidate.model_name,
                            candidate.model_normalized,
                        ),
                    )
                    modelo_insertado = cursor.fetchone()
                else:
                    modelo_insertado = None
                if modelo_insertado is not None:
                    modelo_id = modelo_insertado[0]
                else:
                    cursor.execute(
                        """
                        SELECT id
                        FROM modelos_vehiculos
                        WHERE marca_id = %s
                          AND tipo_vehiculo = %s
                          AND nombre_normalizado = %s
                        """,
                        (
                            marca_id,
                            candidate.tipo_vehiculo,
                            candidate.model_normalized,
                        ),
                    )
                    modelo_existente = cursor.fetchone()
                    if modelo_existente is None:
                        append_error(
                            staged_by_source[candidate.source],
                            "CATALOGO_MODELO_REQUIERE_CASA_CENTRAL",
                        )
                        quarantined_units += 1
                        continue
                    modelo_id = modelo_existente[0]
                version_id = deterministic_id(
                    "version",
                    f"{modelo_id}:{organizacion_id}:unspecified",
                )
                cursor.execute(
                    """
                    INSERT INTO versiones_vehiculos (
                      id,
                      modelo_id,
                      nombre,
                      nombre_normalizado,
                      es_marcador,
                      alcance,
                      organizacion_propietaria_id
                    )
                    VALUES (
                      %s, %s, 'SIN ESPECIFICAR', 'sin especificar', true,
                      'RESTRINGIDO', %s
                    )
                    ON CONFLICT (
                      modelo_id,
                      nombre_normalizado,
                      organizacion_propietaria_id
                    )
                    WHERE alcance = 'RESTRINGIDO'
                    DO UPDATE
                    SET es_marcador = true, actualizado_en = now()
                    RETURNING id
                    """,
                    (version_id, modelo_id, organizacion_id),
                )
                version_id = cursor.fetchone()[0]
                cursor.execute(
                    """
                    INSERT INTO catalogo_organizaciones (
                      organizacion_id, version_id, puede_vender
                    )
                    VALUES (%s, %s, true)
                    ON CONFLICT (organizacion_id, version_id) DO UPDATE
                    SET puede_vender = true, actualizado_en = now()
                    """,
                    (organizacion_id, version_id),
                )
                catalog_cache[catalog_key] = version_id
                catalog_count += 1

            mark_target(
                staged_by_source,
                candidate.source,
                "versiones_vehiculos",
                version_id,
            )

            row_errors = staged_by_source[candidate.source].errors
            row_errors.extend(
                error for error in candidate.errors if error not in row_errors
            )
            if sucursal_id is None:
                row_errors.append("VEHICULO_SUCURSAL_REQUERIDA")
                quarantined_units += 1
                continue
            blocking_errors = {
                "MOTOR_NUMERO_INVALIDO",
                "VIN_COLUMNAS_CONFLICTO",
                "VIN_FALTANTE",
                "VIN_MARCADOR",
                "VIN_LONGITUD_INVALIDA",
                "FECHA_COMPRA_INVALIDA",
            }
            if any(error in blocking_errors for error in candidate.errors):
                quarantined_units += 1
                continue

            unit_id = deterministic_id(
                "vehicle-unit",
                f"{organizacion_id}:{codigo_organizacion}:{candidate.vin_normalizado or ''}",
            )
            proveedor_id = (
                supplier_ids.get(candidate.supplier_normalized)
                if candidate.supplier_normalized
                else None
            )
            if candidate.engine_normalized:
                cursor.execute(
                    """
                    SELECT vin_normalizado
                    FROM unidades_vehiculos
                    WHERE motor_normalizado = %s
                    """,
                    (candidate.engine_normalized,),
                )
                engine_owner = cursor.fetchone()
                if (
                    engine_owner is not None
                    and engine_owner[0] != candidate.vin_normalizado
                ):
                    row_errors.append("MOTOR_NUMERO_CONFLICTO")
                    quarantined_units += 1
                    continue

            fila_importacion_id = staged_by_source[
                candidate.source
            ].fila_importacion_id
            if datos_prueba and fila_importacion_id is None:
                raise RuntimeError("La unidad no tiene fila de importacion vinculada.")

            if datos_prueba:
                sql_unidad = """
                    INSERT INTO unidades_vehiculos (
                      id,
                      organizacion_id,
                      version_id,
                      condicion,
                      vin_mostrado,
                      vin_normalizado,
                      numero_motor,
                      motor_normalizado,
                      anio_fabricacion,
                      kilometraje_km,
                      color,
                      sucursal_id,
                      proveedor_id,
                      origen_adquisicion,
                      costo_compra,
                      estado_inventario,
                      recibido_en,
                      es_importada,
                      fila_importacion_id,
                      datos_inferidos
                    )
                    VALUES (
                      %s, %s, %s, 'NUEVO', %s, %s, %s, %s, %s, 0, %s, %s, %s,
                      'PROVEEDOR', %s, %s, %s, true, %s, %s
                    )
                    ON CONFLICT (vin_normalizado) DO NOTHING
                    RETURNING id
                """
                parametros_unidad = (
                    unit_id,
                    organizacion_id,
                    version_id,
                    candidate.vin_mostrado,
                    candidate.vin_normalizado,
                    candidate.numero_motor,
                    candidate.engine_normalized,
                    candidate.anio_fabricacion,
                    candidate.color,
                    sucursal_id,
                    proveedor_id,
                    candidate.costo_compra,
                    candidate.estado_inventario,
                    candidate.recibido_en,
                    fila_importacion_id,
                    Jsonb(
                        inferred_data(
                            origen="excel_luma",
                            estado_inventario_derivado=candidate.estado_inventario,
                            estado_inventario_sin_confirmar=(
                                candidate.estado_inventario == "BLOQUEADO"
                            ),
                            tipo_vehiculo_inferido_de_ventas=(
                                "VEHICULO_TIPO_INFERIDO_DE_VENTAS"
                                in candidate.errors
                            ),
                            incidencias_origen=sorted(candidate.errors),
                        )
                    ),
                )
            else:
                sql_unidad = """
                    INSERT INTO unidades_vehiculos (
                      id,
                      organizacion_id,
                      version_id,
                      condicion,
                      vin_mostrado,
                      vin_normalizado,
                      numero_motor,
                      motor_normalizado,
                      anio_fabricacion,
                      kilometraje_km,
                      color,
                      sucursal_id,
                      proveedor_id,
                      origen_adquisicion,
                      costo_compra,
                      estado_inventario,
                      recibido_en
                    )
                    VALUES (
                      %s, %s, %s, 'NUEVO', %s, %s, %s, %s, %s, 0, %s, %s, %s,
                      'PROVEEDOR', %s, %s, %s
                    )
                    ON CONFLICT (vin_normalizado) DO NOTHING
                    RETURNING id
                """
                parametros_unidad = (
                    unit_id,
                    organizacion_id,
                    version_id,
                    candidate.vin_mostrado,
                    candidate.vin_normalizado,
                    candidate.numero_motor,
                    candidate.engine_normalized,
                    candidate.anio_fabricacion,
                    candidate.color,
                    sucursal_id,
                    proveedor_id,
                    candidate.costo_compra,
                    candidate.estado_inventario,
                    candidate.recibido_en,
                )

            try:
                with connection.transaction():
                    cursor.execute(sql_unidad, parametros_unidad)
                    inserted = cursor.fetchone()
            except psycopg.errors.UniqueViolation:
                row_errors.append("VEHICULO_UNICIDAD_CONFLICTO")
                quarantined_units += 1
                continue

            if inserted is None:
                cursor.execute(
                    """
                    SELECT
                      id,
                      version_id,
                      sucursal_id,
                      motor_normalizado,
                      anio_fabricacion,
                      color,
                      proveedor_id,
                      costo_compra,
                      recibido_en,
                      condicion
                    FROM unidades_vehiculos
                    WHERE vin_normalizado = %s
                      AND organizacion_id = %s
                    """,
                    (candidate.vin_normalizado, organizacion_id),
                )
                existing = cursor.fetchone()
                if existing is None:
                    row_errors.append("VIN_PERTENECE_OTRA_ORGANIZACION")
                    quarantined_units += 1
                    continue
                (
                    existing_id,
                    existing_version_id,
                    existing_sucursal_id,
                    existing_engine,
                    existing_year,
                    existing_color,
                    existing_supplier_id,
                    existing_purchase_cost,
                    existing_received_at,
                    existing_condition,
                ) = existing
                immutable_conflict = (
                    existing_version_id != version_id
                    or existing_sucursal_id != sucursal_id
                    or existing_engine != candidate.engine_normalized
                    or existing_year != candidate.anio_fabricacion
                    or normalize_text(existing_color) != normalize_text(candidate.color)
                    or existing_supplier_id != proveedor_id
                    or existing_purchase_cost != candidate.costo_compra
                    or existing_received_at != candidate.recibido_en
                    or existing_condition != "NUEVO"
                )
                if immutable_conflict:
                    append_error(
                        staged_by_source[candidate.source],
                        (
                            "VEHICULO_SUCURSAL_CONFLICTO"
                            if existing_sucursal_id != sucursal_id
                            else "VIN_ATRIBUTOS_CONFLICTO"
                        ),
                    )
                    quarantined_units += 1
                    continue
                unit_id = existing_id
            else:
                unit_id = inserted[0]

            mark_target(staged_by_source, candidate.source, "unidades_vehiculos", unit_id)
            unit_count += 1

    return catalog_count, unit_count, quarantined_units


def ejecutar_lote(
    cursor: psycopg.Cursor[Any],
    statement: str,
    parametros: list[tuple[Any, ...]],
) -> None:
    for inicio in range(0, len(parametros), TAMANO_LOTE):
        cursor.executemany(
            statement,
            parametros[inicio : inicio + TAMANO_LOTE],
        )


def resolver_sucursal(
    connection: psycopg.Connection[Any],
    organizacion_id: uuid.UUID,
    nombre_sucursal: str,
) -> uuid.UUID:
    with connection.cursor() as cursor:
        cursor.execute(
            """
            SELECT id
            FROM sucursales
            WHERE organizacion_id = %s
              AND lower(trim(nombre)) = lower(trim(%s))
              AND activa
            """,
            (organizacion_id, nombre_sucursal),
        )
        row = cursor.fetchone()
    if row is None:
        raise RuntimeError(f"Sucursal activa no encontrada: {nombre_sucursal}")
    return row[0]


def asegurar_actor_sistema_importado(
    connection: psycopg.Connection[Any],
    organizacion_id: uuid.UUID,
    sucursal_id: uuid.UUID,
) -> uuid.UUID:
    actor_id = deterministic_id("actor-sistema-importacion", str(organizacion_id))
    with connection.cursor() as cursor:
        cursor.execute(
            """
            INSERT INTO personal (
              id,
              organizacion_id,
              nombre_completo,
              nombre_normalizado,
              sucursal_principal_id,
              puede_iniciar_sesion,
              estado,
              es_actor_sistema_importado,
              datos_inferidos
            )
            VALUES (
              %s, %s, 'IMPORTACION HISTORICA', 'importacion historica', %s,
              false, 'ACTIVO', true, %s
            )
            ON CONFLICT (id) DO UPDATE SET
              es_actor_sistema_importado = true,
              sucursal_principal_id = EXCLUDED.sucursal_principal_id,
              puede_iniciar_sesion = false,
              datos_inferidos = EXCLUDED.datos_inferidos,
              actualizado_en = now()
            """,
            (
                actor_id,
                organizacion_id,
                sucursal_id,
                Jsonb(
                    {
                        "origen": "importacion_excel",
                        "tipo": "actor_sistema_no_personal",
                    }
                ),
            ),
        )
    return actor_id


def cargar_clientes_por_documento(
    connection: psycopg.Connection[Any],
    organizacion_id: uuid.UUID,
) -> dict[tuple[str, str], uuid.UUID]:
    with connection.cursor() as cursor:
        cursor.execute(
            """
            SELECT id, tipo_documento::text, documento_normalizado
            FROM clientes
            WHERE organizacion_id = %s
              AND documento_normalizado IS NOT NULL
            """,
            (organizacion_id,),
        )
        return {
            (str(tipo), documento): cliente_id
            for cliente_id, tipo, documento in cursor.fetchall()
        }


def cargar_unidades_por_vin(
    connection: psycopg.Connection[Any],
    organizacion_id: uuid.UUID,
) -> dict[str, tuple[uuid.UUID, uuid.UUID, str]]:
    with connection.cursor() as cursor:
        cursor.execute(
            """
            SELECT id, vin_normalizado, version_id, condicion::text
            FROM unidades_vehiculos
            WHERE organizacion_id = %s
            """,
            (organizacion_id,),
        )
        return {
            vin: (unidad_id, version_id, str(condicion))
            for unidad_id, vin, version_id, condicion in cursor.fetchall()
        }


def cargar_personal_por_nombre(
    connection: psycopg.Connection[Any],
    organizacion_id: uuid.UUID,
) -> dict[str, uuid.UUID]:
    encontrados: dict[str, list[uuid.UUID]] = defaultdict(list)
    with connection.cursor() as cursor:
        cursor.execute(
            """
            SELECT id, nombre_normalizado
            FROM personal
            WHERE organizacion_id = %s
              AND NOT es_actor_sistema_importado
            """,
            (organizacion_id,),
        )
        for personal_id, nombre_normalizado in cursor.fetchall():
            encontrados[nombre_normalizado].append(personal_id)
    return {
        nombre: ids[0]
        for nombre, ids in encontrados.items()
        if len(ids) == 1
    }


def codigo_cuenta_historica(
    organizacion_id: uuid.UUID, clave: str
) -> str:
    digest = hashlib.sha256(
        f"{organizacion_id}:{clave}".encode("utf-8")
    ).hexdigest()[:24].upper()
    return f"HIST-{digest}"


def clave_cuenta_historica(valor: Any) -> str:
    return normalize_text(valor) or "sin_identificar"


def asegurar_cuentas_historicas(
    connection: psycopg.Connection[Any],
    organizacion_id: uuid.UUID,
    sucursal_id: uuid.UUID,
    valores: dict[str, str | None],
    personal_por_nombre: dict[str, uuid.UUID],
) -> dict[str, uuid.UUID]:
    if not valores:
        return {}
    codigos = {
        clave: codigo_cuenta_historica(organizacion_id, clave)
        for clave in valores
    }
    parametros: list[tuple[Any, ...]] = []
    for clave, mostrado in valores.items():
        cuenta_id = deterministic_id(
            "cuenta-historica", f"{organizacion_id}:{clave}"
        )
        responsable_id = personal_por_nombre.get(clave)
        nombre = (
            f"Cuenta historica importada: {limited_text(mostrado, 100)}"
            if mostrado
            else "Cuenta historica importada sin identificar"
        )
        parametros.append(
            (
                cuenta_id,
                organizacion_id,
                codigos[clave],
                nombre,
                sucursal_id,
                responsable_id,
                Jsonb(
                    inferred_data(
                        origen="excel_luma",
                        clave_historica=clave,
                        responsable_original=mostrado,
                    )
                ),
            )
        )

    with connection.cursor() as cursor:
        ejecutar_lote(
            cursor,
            """
            INSERT INTO cuentas_caja (
              id,
              organizacion_id,
              codigo,
              nombre,
              tipo_cuenta,
              sucursal_id,
              personal_responsable_id,
              activo,
              es_importada,
              datos_inferidos
            )
            VALUES (%s, %s, %s, %s, 'CAJA', %s, %s, true, true, %s)
            ON CONFLICT (organizacion_id, codigo) DO UPDATE SET
              nombre = EXCLUDED.nombre,
              sucursal_id = EXCLUDED.sucursal_id,
              personal_responsable_id = EXCLUDED.personal_responsable_id,
              activo = true,
              es_importada = true,
              datos_inferidos = EXCLUDED.datos_inferidos,
              actualizado_en = now()
            """,
            parametros,
        )
        cursor.execute(
            """
            SELECT id, codigo
            FROM cuentas_caja
            WHERE organizacion_id = %s
              AND codigo = ANY(%s)
            """,
            (organizacion_id, list(codigos.values())),
        )
        ids_por_codigo = {
            codigo: cuenta_id for cuenta_id, codigo in cursor.fetchall()
        }
    return {
        clave: ids_por_codigo[codigo]
        for clave, codigo in codigos.items()
        if codigo in ids_por_codigo
    }


def cargar_operaciones_importadas(
    connection: psycopg.Connection[Any],
    organizacion_id: uuid.UUID,
) -> tuple[
    dict[uuid.UUID, uuid.UUID],
    dict[tuple[uuid.UUID, uuid.UUID], list[uuid.UUID]],
    dict[uuid.UUID, list[uuid.UUID]],
]:
    por_fila: dict[uuid.UUID, uuid.UUID] = {}
    por_cliente_unidad: dict[tuple[uuid.UUID, uuid.UUID], list[uuid.UUID]] = (
        defaultdict(list)
    )
    por_unidad: dict[uuid.UUID, list[uuid.UUID]] = defaultdict(list)
    with connection.cursor() as cursor:
        cursor.execute(
            """
            SELECT id, fila_importacion_id, cliente_id, unidad_vehiculo_id
            FROM operaciones
            WHERE organizacion_id = %s
              AND es_importada
            """,
            (organizacion_id,),
        )
        for operacion_id, fila_id, cliente_id, unidad_id in cursor.fetchall():
            if fila_id is not None:
                por_fila[fila_id] = operacion_id
            if cliente_id is not None and unidad_id is not None:
                por_cliente_unidad[(cliente_id, unidad_id)].append(operacion_id)
                por_unidad[unidad_id].append(operacion_id)
    return por_fila, por_cliente_unidad, por_unidad


def cargar_operaciones_activas_por_unidad(
    connection: psycopg.Connection[Any],
    organizacion_id: uuid.UUID,
) -> dict[uuid.UUID, tuple[uuid.UUID, uuid.UUID | None]]:
    with connection.cursor() as cursor:
        cursor.execute(
            """
            SELECT id, unidad_vehiculo_id, fila_importacion_id
            FROM operaciones
            WHERE organizacion_id = %s
              AND unidad_vehiculo_id IS NOT NULL
              AND estado_operacion IN ('APROBADA', 'CERRADA')
            """,
            (organizacion_id,),
        )
        return {
            unidad_id: (operacion_id, fila_id)
            for operacion_id, unidad_id, fila_id in cursor.fetchall()
        }


def clasificar_venta_por_documento(
    workbook: openpyxl.Workbook,
    venta: StagedRow,
    clientes_por_documento: dict[str, list[StagedRow]],
) -> StagedRow | None:
    documento = normalize_document(cell(workbook, "VENTAS", venta.source.row, 6))
    if documento is None:
        append_error(venta, "VENTA_CLIENTE_DOCUMENTO_INVALIDO")
        return None
    candidatos = clientes_por_documento.get(documento, [])
    if len(candidatos) == 1:
        return candidatos[0]
    if not candidatos:
        append_error(venta, "VENTA_CLIENTE_SIN_COINCIDENCIA")
        return None

    vin_venta, vin_error = valid_vin(cell(workbook, "VENTAS", venta.source.row, 2))
    if vin_error or vin_venta is None:
        append_error(venta, "VENTA_VIN_INVALIDO")
        return None
    por_vin = [
        candidato
        for candidato in candidatos
        if valid_vin(cell(workbook, "Clientes.", candidato.source.row, 15))[0]
        == vin_venta
    ]
    if len(por_vin) == 1:
        return por_vin[0]
    append_error(venta, "VENTA_CLIENTE_DOCUMENTO_AMBIGUO")
    return None


def importar_clientes_canonicos_para_operaciones(
    connection: psycopg.Connection[Any],
    workbook: openpyxl.Workbook,
    staged: list[StagedRow],
    staged_by_source: dict[SourceRef, StagedRow],
    organizacion_id: uuid.UUID,
) -> int:
    """Completa clientes solo cuando la venta eligio su fila sin ambiguedad."""

    clientes_por_documento: dict[str, list[StagedRow]] = defaultdict(list)
    for fila in iter_source_rows(staged, "Clientes."):
        documento = normalize_document(cell(workbook, "Clientes.", fila.source.row, 2))
        if documento:
            clientes_por_documento[documento].append(fila)

    por_documento: dict[str, list[StagedRow]] = defaultdict(list)
    for venta in iter_source_rows(staged, "VENTAS"):
        documento = normalize_document(cell(workbook, "VENTAS", venta.source.row, 6))
        cliente_origen = clasificar_venta_por_documento(
            workbook, venta, clientes_por_documento
        )
        if documento and cliente_origen is not None:
            por_documento[documento].append(cliente_origen)

    existentes = cargar_clientes_por_documento(connection, organizacion_id)
    parametros: list[tuple[Any, ...]] = []
    candidatos_por_documento: dict[str, list[StagedRow]] = {}
    for documento, filas in por_documento.items():
        filas = list({fila.source: fila for fila in filas}.values())
        clave = (tipo_documento(documento), documento)
        if clave in existentes:
            continue
        candidato = min(filas, key=lambda fila: fila.source.row)
        if candidato.fila_importacion_id is None:
            raise RuntimeError("El cliente no tiene fila de importacion vinculada.")
        nombre = safe_text(cell(workbook, "Clientes.", candidato.source.row, 3))
        if nombre is None:
            continue
        cliente_id = deterministic_id(
            "cliente-canonico-importado",
            f"{organizacion_id}:{clave[0]}:{documento}",
        )
        parametros.append(
            (
                cliente_id,
                organizacion_id,
                clave[0],
                documento,
                documento,
                nombre,
                normalize_text(nombre),
                limited_text(
                    cell(workbook, "Clientes.", candidato.source.row, 8), 40
                ),
                candidato.fila_importacion_id,
                Jsonb(
                    inferred_data(
                        origen="excel_luma",
                        seleccion_canonica="fila_menor_de_clientes",
                        filas_clientes_mismo_documento=len(filas),
                    )
                ),
            )
        )
        candidatos_por_documento[documento] = filas

    with connection.cursor() as cursor:
        ejecutar_lote(
            cursor,
            """
            INSERT INTO clientes (
              id,
              organizacion_id,
              tipo_documento,
              numero_documento,
              documento_normalizado,
              nombre_completo,
              nombre_normalizado,
              telefono,
              activo,
              es_importado,
              fila_importacion_id,
              datos_inferidos
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, true, true, %s, %s)
            ON CONFLICT (organizacion_id, tipo_documento, documento_normalizado)
              WHERE documento_normalizado IS NOT NULL
            DO UPDATE SET
              telefono = coalesce(clientes.telefono, EXCLUDED.telefono),
              actualizado_en = now()
            """,
            parametros,
        )
    clientes_reales = cargar_clientes_por_documento(connection, organizacion_id)

    insertados = 0
    for documento, filas in candidatos_por_documento.items():
        cliente_id = clientes_reales.get((tipo_documento(documento), documento))
        if cliente_id is None:
            for fila in filas:
                append_error(fila, "CLIENTE_CANONICO_NO_PERSISTIDO")
            continue
        insertados += 1
        for fila in filas:
            mark_target(staged_by_source, fila.source, "clientes", cliente_id)
    return insertados


def importar_operaciones_historicas(
    connection: psycopg.Connection[Any],
    workbook: openpyxl.Workbook,
    staged: list[StagedRow],
    staged_by_source: dict[SourceRef, StagedRow],
    organizacion_id: uuid.UUID,
    sucursal_id: uuid.UUID,
    actor_sistema_id: uuid.UUID,
) -> dict[str, int]:
    clientes_por_documento: dict[str, list[StagedRow]] = defaultdict(list)
    for fila in iter_source_rows(staged, "Clientes."):
        documento = normalize_document(cell(workbook, "Clientes.", fila.source.row, 2))
        if documento:
            clientes_por_documento[documento].append(fila)

    ids_clientes = cargar_clientes_por_documento(connection, organizacion_id)
    unidades = cargar_unidades_por_vin(connection, organizacion_id)
    personal_por_nombre = cargar_personal_por_nombre(connection, organizacion_id)
    operaciones_activas = cargar_operaciones_activas_por_unidad(
        connection, organizacion_id
    )
    operaciones: list[tuple[Any, ...]] = []
    componentes: dict[uuid.UUID, tuple[uuid.UUID, Decimal, Jsonb]] = {}
    asignaciones: dict[uuid.UUID, tuple[uuid.UUID, str]] = {}
    filas_operacion: list[tuple[StagedRow, StagedRow, uuid.UUID, uuid.UUID]] = []
    clientes_reconciliados: set[SourceRef] = set()

    for venta in iter_source_rows(staged, "VENTAS"):
        fila_importacion_id = venta.fila_importacion_id
        if fila_importacion_id is None:
            raise RuntimeError("La venta no tiene fila de importacion vinculada.")
        fecha = parse_date(cell(workbook, "VENTAS", venta.source.row, 1))
        vin, vin_error = valid_vin(cell(workbook, "VENTAS", venta.source.row, 2))
        documento = normalize_document(cell(workbook, "VENTAS", venta.source.row, 6))
        importe = positive_amount(cell(workbook, "VENTAS", venta.source.row, 9))
        cliente_origen = clasificar_venta_por_documento(
            workbook, venta, clientes_por_documento
        )

        if fecha is None:
            append_error(venta, "VENTA_FECHA_INVALIDA")
        if vin_error or vin is None:
            append_error(venta, "VENTA_VIN_INVALIDO")
        if importe is None:
            append_error(venta, "VENTA_PRECIO_INVALIDO")
        if cliente_origen is None or documento is None or vin is None or importe is None:
            continue

        cliente_id = ids_clientes.get((tipo_documento(documento), documento))
        if cliente_id is None:
            append_error(venta, "VENTA_CLIENTE_NO_IMPORTABLE")
            continue
        unidad = unidades.get(vin)
        if unidad is None:
            append_error(venta, "VENTA_UNIDAD_NO_IMPORTABLE")
            continue
        unidad_id, version_id, condicion = unidad
        operacion_activa = operaciones_activas.get(unidad_id)
        if (
            operacion_activa is not None
            and operacion_activa[1] != fila_importacion_id
        ):
            append_error(venta, "VENTA_UNIDAD_OPERACION_CONFLICTO")
            continue
        vendedor_venta = safe_text(cell(workbook, "VENTAS", venta.source.row, 14))
        vendedor_cliente = safe_text(
            cell(workbook, "Clientes.", cliente_origen.source.row, 12)
        )
        vendedor_para_asignar = vendedor_venta or vendedor_cliente
        vendedor_id = (
            personal_por_nombre.get(normalize_text(vendedor_para_asignar))
            if vendedor_para_asignar
            else None
        )
        operacion_id = deterministic_id(
            "operacion-venta", f"{organizacion_id}:{fila_importacion_id}"
        )
        componente_id = deterministic_id(
            "componente-pago-venta", f"{organizacion_id}:{fila_importacion_id}"
        )
        datos = inferred_data(
            origen="excel_luma",
            precios_referencia="no_disponibles_en_origen",
            estado_operacion_inferido="CERRADA",
            estado_entrega_inferido="NO_PROGRAMADA",
            estado_documentacion_inferido="NO_INICIADA",
            fila_clientes_importacion_id=cliente_origen.fila_importacion_id,
            plataforma_pago_original=cell(
                workbook, "Clientes.", cliente_origen.source.row, 6
            ),
            monto_credito_original=cell(
                workbook, "Clientes.", cliente_origen.source.row, 7
            ),
            respaldo_original=cell(
                workbook, "Clientes.", cliente_origen.source.row, 9
            ),
            entrega_moto_original=cell(
                workbook, "Clientes.", cliente_origen.source.row, 10
            ),
            entrega_papeles_original=cell(
                workbook, "Clientes.", cliente_origen.source.row, 11
            ),
            vendedor_original=vendedor_venta,
            vendedor_clientes_original=vendedor_cliente,
            vendedor_asignado_deterministicamente=vendedor_id is not None,
            origen_vendedor_asignado=(
                "VENTAS" if vendedor_venta else "Clientes."
            )
            if vendedor_id is not None
            else None,
        )
        operaciones.append(
            (
                operacion_id,
                organizacion_id,
                sucursal_id,
                cliente_id,
                version_id,
                condicion,
                unidad_id,
                fecha,
                importe,
                actor_sistema_id,
                fila_importacion_id,
                Jsonb(datos),
            )
        )
        componentes[fila_importacion_id] = (
            componente_id,
            importe,
            Jsonb(
                inferred_data(
                    origen="excel_luma",
                    composicion_detallada_no_comprobable=True,
                    plataforma_pago_original=cell(
                        workbook, "Clientes.", cliente_origen.source.row, 6
                    ),
                    monto_credito_original=cell(
                        workbook, "Clientes.", cliente_origen.source.row, 7
                    ),
                )
            ),
        )
        if vendedor_id is not None:
            asignaciones[operacion_id] = (vendedor_id, "VENDEDOR")
        filas_operacion.append(
            (venta, cliente_origen, operacion_id, componente_id)
        )
        clientes_reconciliados.add(cliente_origen.source)

    with connection.cursor() as cursor:
        ejecutar_lote(
            cursor,
            """
            INSERT INTO operaciones (
              id,
              organizacion_id,
              sucursal_id,
              cliente_id,
              version_id,
              condicion,
              unidad_vehiculo_id,
              fecha_operacion,
              estado_operacion,
              precio_lista,
              precio_minimo,
              precio_acordado,
              moneda,
              estado_entrega,
              estado_documentacion,
              creado_por_personal_id,
              notas,
              es_importada,
              precios_referencia_completos,
              fila_importacion_id,
              datos_inferidos
            )
            VALUES (
              %s, %s, %s, %s, %s, %s, %s, %s, 'CERRADA', NULL, NULL, %s,
              'ARS', 'NO_PROGRAMADA', 'NO_INICIADA', %s,
              'Operacion historica importada; ver datos_inferidos.',
              true, false, %s, %s
            )
            ON CONFLICT (organizacion_id, fila_importacion_id)
              WHERE fila_importacion_id IS NOT NULL
            DO UPDATE SET
              sucursal_id = EXCLUDED.sucursal_id,
              cliente_id = EXCLUDED.cliente_id,
              version_id = EXCLUDED.version_id,
              condicion = EXCLUDED.condicion,
              unidad_vehiculo_id = EXCLUDED.unidad_vehiculo_id,
              fecha_operacion = EXCLUDED.fecha_operacion,
              precio_acordado = EXCLUDED.precio_acordado,
              datos_inferidos = EXCLUDED.datos_inferidos,
              actualizado_en = now()
            """,
            operaciones,
        )
        operaciones_reales: dict[uuid.UUID, uuid.UUID] = {}
        if filas_operacion:
            cursor.execute(
                """
                SELECT id, fila_importacion_id
                FROM operaciones
                WHERE organizacion_id = %s
                  AND fila_importacion_id = ANY(%s)
                """,
                (
                    organizacion_id,
                    [
                        venta.fila_importacion_id
                        for venta, _, _, _ in filas_operacion
                    ],
                ),
            )
            operaciones_reales = {
                fila_id: operacion_id
                for operacion_id, fila_id in cursor.fetchall()
            }
        parametros_componentes = [
            (
                componente_id,
                organizacion_id,
                operaciones_reales[fila_importacion_id],
                importe,
                fila_importacion_id,
                datos,
            )
            for fila_importacion_id, (componente_id, importe, datos)
            in componentes.items()
            if fila_importacion_id in operaciones_reales
        ]
        ejecutar_lote(
            cursor,
            """
            INSERT INTO componentes_pago_operacion (
              id,
              organizacion_id,
              operacion_id,
              tipo_componente,
              importe_esperado,
              estado_pago,
              notas,
              es_importado,
              fila_importacion_id,
              datos_inferidos
            )
            VALUES (
              %s, %s, %s, 'OTRO', %s, 'PENDIENTE',
              'Componente historico sin composicion comprobable; ver datos_inferidos.',
              true, %s, %s
            )
            ON CONFLICT (organizacion_id, fila_importacion_id)
              WHERE fila_importacion_id IS NOT NULL
            DO UPDATE SET
              operacion_id = EXCLUDED.operacion_id,
              importe_esperado = EXCLUDED.importe_esperado,
              datos_inferidos = EXCLUDED.datos_inferidos,
              actualizado_en = now()
            """,
            parametros_componentes,
        )
        componentes_reales: dict[uuid.UUID, uuid.UUID] = {}
        if parametros_componentes:
            cursor.execute(
                """
                SELECT id, fila_importacion_id
                FROM componentes_pago_operacion
                WHERE organizacion_id = %s
                  AND fila_importacion_id = ANY(%s)
                """,
                (
                    organizacion_id,
                    [parametro[-2] for parametro in parametros_componentes],
                ),
            )
            componentes_reales = {
                fila_id: componente_id
                for componente_id, fila_id in cursor.fetchall()
            }
        asignaciones_reales = []
        for venta, _, operacion_id, _ in filas_operacion:
            asignacion = asignaciones.get(operacion_id)
            operacion_real_id = operaciones_reales.get(venta.fila_importacion_id)
            if asignacion is not None and operacion_real_id is not None:
                personal_id, rol = asignacion
                asignaciones_reales.append(
                    (operacion_real_id, organizacion_id, personal_id, rol)
                )
        ejecutar_lote(
            cursor,
            """
            INSERT INTO asignaciones_personal_operacion (
              operacion_id,
              organizacion_id,
              personal_id,
              rol_asignacion
            )
            VALUES (%s, %s, %s, %s)
            ON CONFLICT (operacion_id, personal_id, rol_asignacion) DO NOTHING
            """,
            asignaciones_reales,
        )

    for venta, cliente_origen, _, _ in filas_operacion:
        operacion_real_id = operaciones_reales.get(venta.fila_importacion_id)
        if operacion_real_id is None:
            append_error(venta, "VENTA_OPERACION_NO_PERSISTIDA")
            continue
        mark_target(
            staged_by_source, venta.source, "operaciones", operacion_real_id
        )
        componente_real_id = componentes_reales.get(venta.fila_importacion_id)
        if componente_real_id is None:
            append_error(venta, "VENTA_COMPONENTE_NO_PERSISTIDO")
            continue
        mark_target(
            staged_by_source,
            venta.source,
            "componentes_pago_operacion",
            componente_real_id,
        )
        mark_target(
            staged_by_source,
            cliente_origen.source,
            "operaciones",
            operacion_real_id,
        )

    for cliente in iter_source_rows(staged, "Clientes."):
        if cliente.source not in clientes_reconciliados:
            append_error(cliente, "CLIENTE_VENTA_NO_RECONCILIADA")

    return {
        "operaciones": len(operaciones_reales),
        "componentes_pago": len(componentes_reales),
        "ventas_cuarentena": sum(
            bool(fila.errors) for fila in iter_source_rows(staged, "VENTAS")
        ),
    }


def cuentas_historicas_requeridas(
    workbook: openpyxl.Workbook,
    staged: list[StagedRow],
) -> dict[str, str | None]:
    cuentas: dict[str, str | None] = {}
    for fila in iter_source_rows(staged, "INGRESOS"):
        if es_transferencia_origen(
            cell(workbook, "INGRESOS", fila.source.row, 3)
        ):
            continue
        if (
            parse_date(cell(workbook, "INGRESOS", fila.source.row, 2))
            and positive_amount(cell(workbook, "INGRESOS", fila.source.row, 7))
        ):
            cobrado_por = safe_text(
                cell(workbook, "INGRESOS", fila.source.row, 6)
            )
            cuentas.setdefault(clave_cuenta_historica(cobrado_por), cobrado_por)

    for fila in iter_source_rows(staged, "GASTOS V"):
        importe = positive_amount(cell(workbook, "GASTOS V", fila.source.row, 5))
        estado = estado_pago_desde_origen(
            cell(workbook, "GASTOS V", fila.source.row, 7)
        )
        if importe and estado == "PAGADO":
            pagador = safe_text(cell(workbook, "GASTOS V", fila.source.row, 6))
            cuentas.setdefault(clave_cuenta_historica(pagador), pagador)

    if any(
        positive_amount(cell(workbook, "PAGOS F", fila.source.row, 7))
        for fila in iter_source_rows(staged, "PAGOS F")
    ):
        cuentas.setdefault("sin_identificar", None)
    return cuentas


def importar_ingresos_historicos(
    connection: psycopg.Connection[Any],
    workbook: openpyxl.Workbook,
    staged: list[StagedRow],
    staged_by_source: dict[SourceRef, StagedRow],
    organizacion_id: uuid.UUID,
    sucursal_id: uuid.UUID,
    actor_sistema_id: uuid.UUID,
    cuentas_historicas: dict[str, uuid.UUID],
) -> dict[str, int]:
    ingresos: list[tuple[Any, ...]] = []
    cuentas_por_fila: dict[uuid.UUID, uuid.UUID | None] = {}
    personal_por_nombre = cargar_personal_por_nombre(connection, organizacion_id)
    movimientos_planeados: dict[
        uuid.UUID, tuple[uuid.UUID, uuid.UUID, Decimal, date, uuid.UUID, Jsonb]
    ] = {}
    filas_validas: list[StagedRow] = []

    for fila in iter_source_rows(staged, "INGRESOS"):
        fila_importacion_id = fila.fila_importacion_id
        if fila_importacion_id is None:
            raise RuntimeError("El ingreso no tiene fila de importacion vinculada.")
        fecha = parse_date(cell(workbook, "INGRESOS", fila.source.row, 2))
        importe = positive_amount(cell(workbook, "INGRESOS", fila.source.row, 7))
        if fecha is None:
            append_error(fila, "INGRESO_FECHA_INVALIDA")
        if importe is None:
            append_error(fila, "INGRESO_IMPORTE_INVALIDO")
        if fecha is None or importe is None:
            continue

        tipo_original = limited_text(
            cell(workbook, "INGRESOS", fila.source.row, 3), 120
        )
        descripcion = safe_text(cell(workbook, "INGRESOS", fila.source.row, 5))
        estado_original = limited_text(
            cell(workbook, "INGRESOS", fila.source.row, 8), 120
        )
        observaciones = safe_text(cell(workbook, "INGRESOS", fila.source.row, 9))
        cobrado_por = safe_text(cell(workbook, "INGRESOS", fila.source.row, 6))
        cobrador_id = (
            personal_por_nombre.get(clave_cuenta_historica(cobrado_por))
            if cobrado_por
            else None
        )
        es_transferencia = es_transferencia_origen(tipo_original)
        requiere_conciliacion = es_transferencia
        if es_transferencia:
            estado_registro = "PENDIENTE_CONCILIACION"
            cuenta_id = None
            append_error(fila, "TRANSFERENCIA_CONCILIACION_PENDIENTE")
        else:
            estado_registro = (
                "COBRADO" if es_estado_cobrado(estado_original) else "PENDIENTE"
            )
            cuenta_id = cuentas_historicas.get(
                clave_cuenta_historica(cobrado_por)
            )
            if cuenta_id is None:
                raise RuntimeError("No se pudo resolver la cuenta historica de ingreso.")
            if estado_registro == "PENDIENTE":
                append_error(
                    fila,
                    (
                        "INGRESO_ESTADO_NO_CONFIRMADO"
                        if estado_original
                        else "INGRESO_ESTADO_FALTANTE"
                    ),
                )

        if tipo_original is None:
            tipo_original = "SIN_TIPO_EN_ORIGEN"
            append_error(fila, "INGRESO_TIPO_FALTANTE")
        if descripcion is None:
            descripcion = "Descripcion no informada en origen"
            append_error(fila, "INGRESO_DESCRIPCION_FALTANTE")
        ingreso_id = deterministic_id(
            "ingreso-historico", f"{organizacion_id}:{fila_importacion_id}"
        )
        datos = inferred_data(
            origen="excel_luma",
            vin_original=cell(workbook, "INGRESOS", fila.source.row, 4),
            tipo_original_completo=cell(workbook, "INGRESOS", fila.source.row, 3),
            estado_original_completo=cell(
                workbook, "INGRESOS", fila.source.row, 8
            ),
            hora_no_informada=True,
            transferencia_sin_origen_y_destino_comprobados=es_transferencia,
        )
        ingresos.append(
            (
                ingreso_id,
                organizacion_id,
                sucursal_id,
                fila_importacion_id,
                fecha,
                tipo_original,
                descripcion,
                importe,
                estado_registro,
                estado_original,
                observaciones,
                cobrado_por,
                cobrador_id,
                cuenta_id,
                es_transferencia,
                requiere_conciliacion,
                Jsonb(datos),
            )
        )
        cuentas_por_fila[fila_importacion_id] = cuenta_id
        if estado_registro == "COBRADO" and cuenta_id is not None:
            movimiento_id = deterministic_id(
                "movimiento-ingreso-historico",
                f"{organizacion_id}:{fila_importacion_id}",
            )
            movimientos_planeados[fila_importacion_id] = (
                movimiento_id,
                ingreso_id,
                importe,
                fecha,
                cuenta_id,
                Jsonb(
                    inferred_data(
                        origen="excel_luma",
                        hora_no_informada=True,
                        tipo_ingreso_original=tipo_original,
                    )
                ),
            )
        filas_validas.append(fila)

    ingresos_reales: dict[uuid.UUID, uuid.UUID] = {}
    movimientos_reales: dict[uuid.UUID, uuid.UUID] = {}
    with connection.cursor() as cursor:
        ejecutar_lote(
            cursor,
            """
            INSERT INTO ingresos (
              id,
              organizacion_id,
              sucursal_id,
              fila_importacion_id,
              fecha_ingreso,
              tipo_original,
              descripcion,
              importe,
              estado_registro,
              estado_original,
              observaciones,
              cobrado_por_original,
              cobrado_por_personal_id,
              cuenta_caja_id,
              es_transferencia,
              requiere_conciliacion,
              datos_inferidos
            )
            VALUES (
              %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
            )
            ON CONFLICT (organizacion_id, fila_importacion_id) DO UPDATE SET
              sucursal_id = EXCLUDED.sucursal_id,
              fecha_ingreso = EXCLUDED.fecha_ingreso,
              tipo_original = EXCLUDED.tipo_original,
              descripcion = EXCLUDED.descripcion,
              importe = EXCLUDED.importe,
              estado_registro = EXCLUDED.estado_registro,
              estado_original = EXCLUDED.estado_original,
              observaciones = EXCLUDED.observaciones,
              cobrado_por_original = EXCLUDED.cobrado_por_original,
              cobrado_por_personal_id = EXCLUDED.cobrado_por_personal_id,
              cuenta_caja_id = EXCLUDED.cuenta_caja_id,
              es_transferencia = EXCLUDED.es_transferencia,
              requiere_conciliacion = EXCLUDED.requiere_conciliacion,
              datos_inferidos = EXCLUDED.datos_inferidos,
              actualizado_en = now()
            """,
            ingresos,
        )
        if filas_validas:
            cursor.execute(
                """
                SELECT id, fila_importacion_id
                FROM ingresos
                WHERE organizacion_id = %s
                  AND fila_importacion_id = ANY(%s)
                """,
                (
                    organizacion_id,
                    [fila.fila_importacion_id for fila in filas_validas],
                ),
            )
            ingresos_reales = {
                fila_id: ingreso_id
                for ingreso_id, fila_id in cursor.fetchall()
            }
        movimientos = [
            (
                movimiento_id,
                organizacion_id,
                cuenta_id,
                importe,
                fecha_historica(fecha),
                ingresos_reales[fila_id],
                f"INGRESO-HIST-{str(fila_id)[:8].upper()}",
                actor_sistema_id,
                fila_id,
                datos,
            )
            for fila_id, (
                movimiento_id,
                _,
                importe,
                fecha,
                cuenta_id,
                datos,
            ) in movimientos_planeados.items()
            if fila_id in ingresos_reales
        ]
        ejecutar_lote(
            cursor,
            """
            INSERT INTO movimientos_caja (
              id,
              organizacion_id,
              cuenta_caja_id,
              tipo_movimiento,
              direccion,
              importe,
              contabilizado_en,
              ingreso_id,
              referencia,
              notas,
              registrado_por_personal_id,
              es_importado,
              fila_importacion_id,
              datos_inferidos
            )
            VALUES (
              %s, %s, %s, 'INGRESO', 'CREDITO', %s, %s, %s, %s,
              'Movimiento historico importado; hora no informada.',
              %s, true, %s, %s
            )
            ON CONFLICT (organizacion_id, fila_importacion_id)
              WHERE fila_importacion_id IS NOT NULL
            DO UPDATE SET
              cuenta_caja_id = EXCLUDED.cuenta_caja_id,
              importe = EXCLUDED.importe,
              contabilizado_en = EXCLUDED.contabilizado_en,
              ingreso_id = EXCLUDED.ingreso_id,
              datos_inferidos = EXCLUDED.datos_inferidos
            """,
            movimientos,
        )
        if movimientos:
            cursor.execute(
                """
                SELECT id, fila_importacion_id
                FROM movimientos_caja
                WHERE organizacion_id = %s
                  AND fila_importacion_id = ANY(%s)
                """,
                (organizacion_id, list(movimientos_planeados)),
            )
            movimientos_reales = {
                fila_id: movimiento_id
                for movimiento_id, fila_id in cursor.fetchall()
            }

    for fila in filas_validas:
        fila_id = fila.fila_importacion_id
        ingreso_id = ingresos_reales.get(fila_id)
        if ingreso_id is None:
            append_error(fila, "INGRESO_NO_PERSISTIDO")
            continue
        mark_target(staged_by_source, fila.source, "ingresos", ingreso_id)
        cuenta_id = cuentas_por_fila.get(fila_id)
        if cuenta_id is not None:
            mark_target(
                staged_by_source,
                fila.source,
                "cuentas_caja",
                cuenta_id,
            )
        movimiento_id = movimientos_reales.get(fila_id)
        if movimiento_id is not None:
            mark_target(
                staged_by_source, fila.source, "movimientos_caja", movimiento_id
            )

    return {
        "ingresos": len(ingresos_reales),
        "movimientos_ingresos": len(movimientos_reales),
        "transferencias_pendientes": sum(
            "TRANSFERENCIA_CONCILIACION_PENDIENTE" in fila.errors
            for fila in iter_source_rows(staged, "INGRESOS")
        ),
    }


def persistir_gastos_historicos(
    connection: psycopg.Connection[Any],
    gastos: list[GastoHistorico],
    staged_by_source: dict[SourceRef, StagedRow],
    organizacion_id: uuid.UUID,
    sucursal_id: uuid.UUID,
    actor_sistema_id: uuid.UUID,
) -> dict[str, int]:
    por_fila: dict[uuid.UUID, tuple[GastoHistorico, uuid.UUID]] = {}
    parametros_gastos: list[tuple[Any, ...]] = []
    for gasto in gastos:
        fila_importacion_id = gasto.source.fila_importacion_id
        if fila_importacion_id is None:
            raise RuntimeError("El gasto no tiene fila de importacion vinculada.")
        gasto_id = deterministic_id(
            "gasto-historico", f"{organizacion_id}:{fila_importacion_id}"
        )
        por_fila[fila_importacion_id] = (gasto, gasto_id)
        parametros_gastos.append(
            (
                gasto_id,
                organizacion_id,
                sucursal_id,
                gasto.unidad_vehiculo_id,
                gasto.categoria,
                gasto.detalle,
                gasto.fecha_generacion,
                gasto.importe,
                gasto.recuperable,
                gasto.estado_pago,
                actor_sistema_id,
                fila_importacion_id,
                gasto.pagador_original,
                gasto.recuperable_original,
                gasto.referencia_origen,
                gasto.vin_origen_mostrado,
                gasto.vin_origen_normalizado,
                Jsonb(gasto.datos_inferidos),
            )
        )

    gastos_reales: dict[uuid.UUID, uuid.UUID] = {}
    movimientos_reales: dict[uuid.UUID, uuid.UUID] = {}
    with connection.cursor() as cursor:
        ejecutar_lote(
            cursor,
            """
            INSERT INTO gastos (
              id,
              organizacion_id,
              sucursal_id,
              unidad_vehiculo_id,
              categoria,
              detalle,
              fecha_generacion,
              importe,
              recuperable,
              estado_pago,
              creado_por_personal_id,
              es_importado,
              fila_importacion_id,
              pagador_original,
              recuperable_original,
              referencia_origen,
              vin_origen_mostrado,
              vin_origen_normalizado,
              datos_inferidos
            )
            VALUES (
              %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, true, %s,
              %s, %s, %s, %s, %s, %s
            )
            ON CONFLICT (organizacion_id, fila_importacion_id)
              WHERE fila_importacion_id IS NOT NULL
            DO UPDATE SET
              sucursal_id = EXCLUDED.sucursal_id,
              unidad_vehiculo_id = EXCLUDED.unidad_vehiculo_id,
              categoria = EXCLUDED.categoria,
              detalle = EXCLUDED.detalle,
              fecha_generacion = EXCLUDED.fecha_generacion,
              importe = EXCLUDED.importe,
              recuperable = EXCLUDED.recuperable,
              estado_pago = EXCLUDED.estado_pago,
              pagador_original = EXCLUDED.pagador_original,
              recuperable_original = EXCLUDED.recuperable_original,
              referencia_origen = EXCLUDED.referencia_origen,
              vin_origen_mostrado = EXCLUDED.vin_origen_mostrado,
              vin_origen_normalizado = EXCLUDED.vin_origen_normalizado,
              datos_inferidos = EXCLUDED.datos_inferidos,
              actualizado_en = now()
            """,
            parametros_gastos,
        )
        if por_fila:
            cursor.execute(
                """
                SELECT id, fila_importacion_id
                FROM gastos
                WHERE organizacion_id = %s
                  AND fila_importacion_id = ANY(%s)
                """,
                (organizacion_id, list(por_fila)),
            )
            gastos_reales = {
                fila_id: gasto_id for gasto_id, fila_id in cursor.fetchall()
            }

        parametros_movimientos: list[tuple[Any, ...]] = []
        for fila_id, (gasto, _) in por_fila.items():
            gasto_real_id = gastos_reales.get(fila_id)
            if (
                gasto_real_id is None
                or gasto.estado_pago != "PAGADO"
                or gasto.cuenta_caja_id is None
            ):
                continue
            movimiento_id = deterministic_id(
                "movimiento-gasto-historico",
                f"{organizacion_id}:{fila_id}",
            )
            parametros_movimientos.append(
                (
                    movimiento_id,
                    organizacion_id,
                    gasto.cuenta_caja_id,
                    gasto.importe,
                    fecha_historica(gasto.fecha_generacion),
                    gasto_real_id,
                    f"GASTO-HIST-{str(fila_id)[:8].upper()}",
                    actor_sistema_id,
                    fila_id,
                    Jsonb(
                        inferred_data(
                            origen="excel_luma",
                            hora_no_informada=True,
                            categoria_original=gasto.categoria,
                        )
                    ),
                )
            )
        ejecutar_lote(
            cursor,
            """
            INSERT INTO movimientos_caja (
              id,
              organizacion_id,
              cuenta_caja_id,
              tipo_movimiento,
              direccion,
              importe,
              contabilizado_en,
              gasto_id,
              referencia,
              notas,
              registrado_por_personal_id,
              es_importado,
              fila_importacion_id,
              datos_inferidos
            )
            VALUES (
              %s, %s, %s, 'EGRESO', 'DEBITO', %s, %s, %s, %s,
              'Movimiento historico importado; hora no informada.',
              %s, true, %s, %s
            )
            ON CONFLICT (organizacion_id, fila_importacion_id)
              WHERE fila_importacion_id IS NOT NULL
            DO UPDATE SET
              cuenta_caja_id = EXCLUDED.cuenta_caja_id,
              importe = EXCLUDED.importe,
              contabilizado_en = EXCLUDED.contabilizado_en,
              gasto_id = EXCLUDED.gasto_id,
              datos_inferidos = EXCLUDED.datos_inferidos
            """,
            parametros_movimientos,
        )
        if parametros_movimientos:
            cursor.execute(
                """
                SELECT id, fila_importacion_id
                FROM movimientos_caja
                WHERE organizacion_id = %s
                  AND fila_importacion_id = ANY(%s)
                """,
                (
                    organizacion_id,
                    [parametro[-2] for parametro in parametros_movimientos],
                ),
            )
            movimientos_reales = {
                fila_id: movimiento_id
                for movimiento_id, fila_id in cursor.fetchall()
            }

    for fila_id, (gasto, _) in por_fila.items():
        gasto_real_id = gastos_reales.get(fila_id)
        if gasto_real_id is None:
            append_error(gasto.source, "GASTO_NO_PERSISTIDO")
            continue
        mark_target(
            staged_by_source, gasto.source.source, "gastos", gasto_real_id
        )
        if gasto.cuenta_caja_id is not None:
            mark_target(
                staged_by_source,
                gasto.source.source,
                "cuentas_caja",
                gasto.cuenta_caja_id,
            )
        movimiento_id = movimientos_reales.get(fila_id)
        if movimiento_id is not None:
            mark_target(
                staged_by_source,
                gasto.source.source,
                "movimientos_caja",
                movimiento_id,
            )
    return {
        "gastos": len(gastos_reales),
        "movimientos_gastos": len(movimientos_reales),
    }


def importar_gastos_ventas_historicos(
    connection: psycopg.Connection[Any],
    workbook: openpyxl.Workbook,
    staged: list[StagedRow],
    staged_by_source: dict[SourceRef, StagedRow],
    organizacion_id: uuid.UUID,
    sucursal_id: uuid.UUID,
    actor_sistema_id: uuid.UUID,
    cuentas_historicas: dict[str, uuid.UUID],
) -> dict[str, int]:
    gastos: list[GastoHistorico] = []
    for fila in iter_source_rows(staged, "GASTOS V"):
        fecha = parse_date(cell(workbook, "GASTOS V", fila.source.row, 2))
        importe = positive_amount(cell(workbook, "GASTOS V", fila.source.row, 5))
        if fecha is None:
            append_error(fila, "GASTO_FECHA_INVALIDA")
        if importe is None:
            append_error(fila, "GASTO_IMPORTE_INVALIDO")
        if fecha is None or importe is None:
            continue
        categoria = limited_text(
            cell(workbook, "GASTOS V", fila.source.row, 3), 100
        )
        detalle = safe_text(cell(workbook, "GASTOS V", fila.source.row, 4))
        pagador = safe_text(cell(workbook, "GASTOS V", fila.source.row, 6))
        estado_original = cell(workbook, "GASTOS V", fila.source.row, 7)
        estado_pago = estado_pago_desde_origen(estado_original)
        recuperable_original = safe_text(
            cell(workbook, "GASTOS V", fila.source.row, 8)
        )
        recuperable = booleano_desde_origen(
            cell(workbook, "GASTOS V", fila.source.row, 8)
        )
        if categoria is None:
            categoria = "SIN_CATEGORIA_EN_ORIGEN"
            append_error(fila, "GASTO_CATEGORIA_FALTANTE")
        if detalle is None:
            detalle = "Detalle no informado en origen"
            append_error(fila, "GASTO_DETALLE_FALTANTE")
        if recuperable is None:
            recuperable = False
            if recuperable_original is not None:
                append_error(fila, "GASTO_RECUPERABLE_NO_INTERPRETABLE")
        if estado_pago == "PENDIENTE":
            append_error(
                fila,
                (
                    "GASTO_ESTADO_NO_CONFIRMADO"
                    if safe_text(estado_original)
                    else "GASTO_ESTADO_FALTANTE"
                ),
            )
        cuenta_id = None
        if estado_pago == "PAGADO":
            cuenta_id = cuentas_historicas.get(clave_cuenta_historica(pagador))
            if cuenta_id is None:
                raise RuntimeError("No se pudo resolver la cuenta historica de gasto.")
        gastos.append(
            GastoHistorico(
                source=fila,
                categoria=categoria,
                detalle=detalle,
                fecha_generacion=fecha,
                importe=importe,
                recuperable=recuperable,
                estado_pago=estado_pago,
                pagador_original=pagador,
                recuperable_original=recuperable_original,
                referencia_origen="GASTOS V",
                vin_origen_mostrado=None,
                vin_origen_normalizado=None,
                unidad_vehiculo_id=None,
                datos_inferidos=inferred_data(
                    origen="excel_luma",
                    estado_original=estado_original,
                    observaciones_original=cell(
                        workbook, "GASTOS V", fila.source.row, 11
                    ),
                    recuperable_interpretado=recuperable,
                    recuperable_no_informado=recuperable_original is None,
                    estado_pago_inferido_desde_origen=estado_pago,
                    hora_no_informada=True,
                ),
                cuenta_caja_id=cuenta_id,
            )
        )
    return persistir_gastos_historicos(
        connection,
        gastos,
        staged_by_source,
        organizacion_id,
        sucursal_id,
        actor_sistema_id,
    )


def importar_pagos_formulario_historicos(
    connection: psycopg.Connection[Any],
    workbook: openpyxl.Workbook,
    staged: list[StagedRow],
    staged_by_source: dict[SourceRef, StagedRow],
    organizacion_id: uuid.UUID,
    sucursal_id: uuid.UUID,
    actor_sistema_id: uuid.UUID,
    cuentas_historicas: dict[str, uuid.UUID],
) -> dict[str, int]:
    unidades = cargar_unidades_por_vin(connection, organizacion_id)
    cuenta_sin_identificar = cuentas_historicas.get("sin_identificar")
    if cuenta_sin_identificar is None:
        raise RuntimeError("No se pudo resolver la cuenta historica sin identificar.")
    gastos: list[GastoHistorico] = []
    for fila in iter_source_rows(staged, "PAGOS F"):
        fecha = parse_date(cell(workbook, "PAGOS F", fila.source.row, 2))
        importe = positive_amount(cell(workbook, "PAGOS F", fila.source.row, 7))
        if fecha is None:
            append_error(fila, "PAGO_F_FECHA_INVALIDA")
        if importe is None:
            append_error(fila, "PAGO_F_IMPORTE_INVALIDO")
        if fecha is None or importe is None:
            continue
        categoria = limited_text(cell(workbook, "PAGOS F", fila.source.row, 3), 100)
        vehiculo = safe_text(cell(workbook, "PAGOS F", fila.source.row, 5))
        proveedor = safe_text(cell(workbook, "PAGOS F", fila.source.row, 6))
        estado_original = cell(workbook, "PAGOS F", fila.source.row, 8)
        estado_pago = estado_pago_desde_origen(estado_original)
        vin_mostrado = limited_text(cell(workbook, "PAGOS F", fila.source.row, 4), 40)
        vin_normalizado, vin_error = valid_vin(vin_mostrado)
        unidad_id = None
        if vin_mostrado:
            if vin_error or vin_normalizado is None:
                append_error(fila, "PAGO_F_VIN_INVALIDO")
            elif vin_normalizado in unidades:
                unidad_id = unidades[vin_normalizado][0]
            else:
                append_error(fila, "PAGO_F_VIN_NO_CONCILIADO")
        if categoria is None:
            categoria = "SIN_CATEGORIA_EN_ORIGEN"
            append_error(fila, "PAGO_F_CATEGORIA_FALTANTE")
        if estado_pago == "PENDIENTE":
            append_error(
                fila,
                (
                    "PAGO_F_ESTADO_NO_CONFIRMADO"
                    if safe_text(estado_original)
                    else "PAGO_F_ESTADO_FALTANTE"
                ),
            )
        detalle_partes = [
            parte
            for parte in (
                vehiculo,
                f"Proveedor: {proveedor}" if proveedor else None,
            )
            if parte
        ]
        detalle = " | ".join(detalle_partes) or "Detalle no informado en origen"
        if not detalle_partes:
            append_error(fila, "PAGO_F_DETALLE_FALTANTE")
        gastos.append(
            GastoHistorico(
                source=fila,
                categoria=categoria,
                detalle=detalle,
                fecha_generacion=fecha,
                importe=importe,
                recuperable=False,
                estado_pago=estado_pago,
                pagador_original=None,
                recuperable_original=None,
                referencia_origen=vehiculo,
                vin_origen_mostrado=vin_mostrado,
                vin_origen_normalizado=vin_normalizado,
                unidad_vehiculo_id=unidad_id,
                datos_inferidos=inferred_data(
                    origen="excel_luma",
                    proveedor_original=proveedor,
                    estado_original=estado_original,
                    referencia_vehiculo_original=vehiculo,
                    hora_no_informada=True,
                ),
                cuenta_caja_id=(
                    cuenta_sin_identificar
                    if estado_pago == "PAGADO"
                    else None
                ),
            )
        )
    return persistir_gastos_historicos(
        connection,
        gastos,
        staged_by_source,
        organizacion_id,
        sucursal_id,
        actor_sistema_id,
    )


def importar_polizas_historicas(
    connection: psycopg.Connection[Any],
    workbook: openpyxl.Workbook,
    staged: list[StagedRow],
    staged_by_source: dict[SourceRef, StagedRow],
    organizacion_id: uuid.UUID,
    sucursal_id: uuid.UUID,
) -> dict[str, int]:
    clientes = cargar_clientes_por_documento(connection, organizacion_id)
    unidades = cargar_unidades_por_vin(connection, organizacion_id)
    _, operaciones_cliente_unidad, operaciones_unidad = (
        cargar_operaciones_importadas(connection, organizacion_id)
    )
    polizas: list[tuple[Any, ...]] = []
    filas_validas: list[StagedRow] = []
    for fila in iter_source_rows(staged, "SEGUROS"):
        fila_importacion_id = fila.fila_importacion_id
        if fila_importacion_id is None:
            raise RuntimeError("El seguro no tiene fila de importacion vinculada.")
        fecha = parse_date(cell(workbook, "SEGUROS", fila.source.row, 3))
        importe = positive_amount(cell(workbook, "SEGUROS", fila.source.row, 6))
        if fecha is None:
            append_error(fila, "SEGURO_FECHA_INVALIDA")
        if importe is None:
            append_error(fila, "SEGURO_IMPORTE_INVALIDO")
        if fecha is None or importe is None:
            continue

        documento_mostrado = limited_text(
            cell(workbook, "SEGUROS", fila.source.row, 1), 30
        )
        documento = normalize_document(documento_mostrado)
        nombre = limited_text(cell(workbook, "SEGUROS", fila.source.row, 2), 180)
        referencia_vehiculo = safe_text(cell(workbook, "SEGUROS", fila.source.row, 4))
        cliente_id = (
            clientes.get((tipo_documento(documento), documento))
            if documento
            else None
        )
        if documento_mostrado and cliente_id is None:
            append_error(fila, "SEGURO_CLIENTE_NO_CONCILIADO")

        vin_referencia = normalize_identifier(referencia_vehiculo)
        unidad = unidades.get(vin_referencia)
        unidad_id = unidad[0] if unidad else None
        if referencia_vehiculo and unidad_id is None:
            posible_vin, _ = valid_vin(referencia_vehiculo)
            if posible_vin:
                append_error(fila, "SEGURO_UNIDAD_NO_CONCILIADA")
        operacion_id = None
        if cliente_id is not None and unidad_id is not None:
            candidatas = operaciones_cliente_unidad.get(
                (cliente_id, unidad_id), []
            )
            if len(candidatas) == 1:
                operacion_id = candidatas[0]
        elif unidad_id is not None:
            candidatas = operaciones_unidad.get(unidad_id, [])
            if len(candidatas) == 1:
                operacion_id = candidatas[0]

        poliza_id = deterministic_id(
            "poliza-historica", f"{organizacion_id}:{fila_importacion_id}"
        )
        polizas.append(
            (
                poliza_id,
                organizacion_id,
                sucursal_id,
                fila_importacion_id,
                cliente_id,
                operacion_id,
                unidad_id,
                fecha,
                safe_text(cell(workbook, "SEGUROS", fila.source.row, 5)),
                documento_mostrado,
                nombre,
                referencia_vehiculo,
                importe,
                Jsonb(
                    inferred_data(
                        origen="excel_luma",
                        documento_normalizado=documento,
                        fecha_vigencia_no_informada=True,
                        estado_vigencia_no_informado=True,
                    )
                ),
            )
        )
        filas_validas.append(fila)

    polizas_reales: dict[uuid.UUID, uuid.UUID] = {}
    with connection.cursor() as cursor:
        ejecutar_lote(
            cursor,
            """
            INSERT INTO polizas_seguros (
              id,
              organizacion_id,
              sucursal_id,
              fila_importacion_id,
              cliente_id,
              operacion_id,
              unidad_vehiculo_id,
              fecha_poliza,
              aseguradora,
              referencia_documento_cliente,
              referencia_nombre_cliente,
              referencia_vehiculo,
              importe,
              datos_inferidos
            )
            VALUES (
              %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
            )
            ON CONFLICT (organizacion_id, fila_importacion_id) DO UPDATE SET
              sucursal_id = EXCLUDED.sucursal_id,
              cliente_id = EXCLUDED.cliente_id,
              operacion_id = EXCLUDED.operacion_id,
              unidad_vehiculo_id = EXCLUDED.unidad_vehiculo_id,
              fecha_poliza = EXCLUDED.fecha_poliza,
              aseguradora = EXCLUDED.aseguradora,
              referencia_documento_cliente = EXCLUDED.referencia_documento_cliente,
              referencia_nombre_cliente = EXCLUDED.referencia_nombre_cliente,
              referencia_vehiculo = EXCLUDED.referencia_vehiculo,
              importe = EXCLUDED.importe,
              datos_inferidos = EXCLUDED.datos_inferidos,
              actualizado_en = now()
            """,
            polizas,
        )
        if filas_validas:
            cursor.execute(
                """
                SELECT id, fila_importacion_id
                FROM polizas_seguros
                WHERE organizacion_id = %s
                  AND fila_importacion_id = ANY(%s)
                """,
                (
                    organizacion_id,
                    [fila.fila_importacion_id for fila in filas_validas],
                ),
            )
            polizas_reales = {
                fila_id: poliza_id
                for poliza_id, fila_id in cursor.fetchall()
            }
    for fila in filas_validas:
        poliza_id = polizas_reales.get(fila.fila_importacion_id)
        if poliza_id is None:
            append_error(fila, "SEGURO_NO_PERSISTIDO")
        else:
            mark_target(staged_by_source, fila.source, "polizas_seguros", poliza_id)
    return {"polizas_seguros": len(polizas_reales)}


def importar_prospectos_historicos(
    connection: psycopg.Connection[Any],
    workbook: openpyxl.Workbook,
    staged: list[StagedRow],
    staged_by_source: dict[SourceRef, StagedRow],
    organizacion_id: uuid.UUID,
    sucursal_id: uuid.UUID,
) -> dict[str, int]:
    bloques = (
        ("bloque_1", 2, 3, 4, 5, 6),
        ("bloque_2", 9, 10, 11, 12, 13),
        ("bloque_3", 16, 17, 18, 19, 20),
    )
    prospectos: list[tuple[Any, ...]] = []
    origenes: list[tuple[StagedRow, str]] = []
    for fila in iter_source_rows(staged, "clientes potenciales."):
        fila_importacion_id = fila.fila_importacion_id
        if fila_importacion_id is None:
            raise RuntimeError("El prospecto no tiene fila de importacion vinculada.")
        for bloque, col_nombre, col_documento, col_telefono, col_estado, col_comision in bloques:
            nombre = limited_text(
                cell(
                    workbook,
                    "clientes potenciales.",
                    fila.source.row,
                    col_nombre,
                ),
                180,
            )
            documento_mostrado = limited_text(
                cell(
                    workbook,
                    "clientes potenciales.",
                    fila.source.row,
                    col_documento,
                ),
                40,
            )
            telefono = limited_text(
                cell(
                    workbook,
                    "clientes potenciales.",
                    fila.source.row,
                    col_telefono,
                ),
                40,
            )
            estado = limited_text(
                cell(
                    workbook,
                    "clientes potenciales.",
                    fila.source.row,
                    col_estado,
                ),
                160,
            )
            comision_original = safe_text(
                cell(
                    workbook,
                    "clientes potenciales.",
                    fila.source.row,
                    col_comision,
                )
            )
            if not any(
                (nombre, documento_mostrado, telefono, estado, comision_original)
            ):
                continue
            if not any((nombre, documento_mostrado, telefono)):
                append_error(fila, "PROSPECTO_IDENTIDAD_FALTANTE")
                continue
            comision = parse_amount(comision_original)
            if comision_original and comision is None:
                append_error(fila, "PROSPECTO_COMISION_INVALIDA")
            prospecto_id = deterministic_id(
                "prospecto-historico",
                f"{organizacion_id}:{fila_importacion_id}:{bloque}",
            )
            prospectos.append(
                (
                    prospecto_id,
                    organizacion_id,
                    sucursal_id,
                    fila_importacion_id,
                    bloque,
                    nombre,
                    documento_mostrado,
                    normalize_document(documento_mostrado),
                    telefono,
                    estado,
                    comision,
                    comision_original,
                    Jsonb(
                        inferred_data(
                            origen="excel_luma",
                            bloque_origen=bloque,
                            documento_original_completo=cell(
                                workbook,
                                "clientes potenciales.",
                                fila.source.row,
                                col_documento,
                            ),
                        )
                    ),
                )
            )
            origenes.append((fila, bloque))

    prospectos_reales: dict[tuple[uuid.UUID, str], uuid.UUID] = {}
    with connection.cursor() as cursor:
        ejecutar_lote(
            cursor,
            """
            INSERT INTO prospectos (
              id,
              organizacion_id,
              sucursal_id,
              fila_importacion_id,
              bloque_origen,
              nombre_completo,
              documento_mostrado,
              documento_normalizado,
              telefono,
              estado_original,
              comision_referido,
              comision_referido_original,
              datos_inferidos
            )
            VALUES (
              %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
            )
            ON CONFLICT (
              organizacion_id, fila_importacion_id, bloque_origen
            )
            DO UPDATE SET
              sucursal_id = EXCLUDED.sucursal_id,
              nombre_completo = EXCLUDED.nombre_completo,
              documento_mostrado = EXCLUDED.documento_mostrado,
              documento_normalizado = EXCLUDED.documento_normalizado,
              telefono = EXCLUDED.telefono,
              estado_original = EXCLUDED.estado_original,
              comision_referido = EXCLUDED.comision_referido,
              comision_referido_original = EXCLUDED.comision_referido_original,
              datos_inferidos = EXCLUDED.datos_inferidos,
              actualizado_en = now()
            """,
            prospectos,
        )
        if origenes:
            cursor.execute(
                """
                SELECT id, fila_importacion_id, bloque_origen
                FROM prospectos
                WHERE organizacion_id = %s
                  AND fila_importacion_id = ANY(%s)
                """,
                (
                    organizacion_id,
                    list({fila.fila_importacion_id for fila, _ in origenes}),
                ),
            )
            prospectos_reales = {
                (fila_id, bloque): prospecto_id
                for prospecto_id, fila_id, bloque in cursor.fetchall()
            }
    for fila, bloque in origenes:
        prospecto_id = prospectos_reales.get((fila.fila_importacion_id, bloque))
        if prospecto_id is None:
            append_error(fila, "PROSPECTO_NO_PERSISTIDO")
        else:
            mark_target(
                staged_by_source,
                fila.source,
                "prospectos",
                prospecto_id,
                bloque,
            )
    return {"prospectos": len(prospectos_reales)}


def importar_registros_inventario_historicos(
    connection: psycopg.Connection[Any],
    workbook: openpyxl.Workbook,
    staged: list[StagedRow],
    staged_by_source: dict[SourceRef, StagedRow],
    organizacion_id: uuid.UUID,
    sucursal_id: uuid.UUID,
) -> dict[str, int]:
    unidades = cargar_unidades_por_vin(connection, organizacion_id)
    registros: list[tuple[Any, ...]] = []
    filas: list[StagedRow] = []
    for hoja in ("SERGIO", "SIAM"):
        for fila in iter_source_rows(staged, hoja):
            fila_importacion_id = fila.fila_importacion_id
            if fila_importacion_id is None:
                raise RuntimeError(
                    "El registro de inventario no tiene fila de importacion vinculada."
                )
            fecha = parse_date(cell(workbook, hoja, fila.source.row, 1))
            vin_mostrado = limited_text(cell(workbook, hoja, fila.source.row, 3), 40)
            vin_normalizado, vin_error = valid_vin(vin_mostrado)
            if fecha is None:
                append_error(fila, "INVENTARIO_FECHA_INVALIDA")
            if vin_mostrado and (vin_error or vin_normalizado is None):
                append_error(fila, "INVENTARIO_VIN_INVALIDO")
            unidad_id = (
                unidades[vin_normalizado][0]
                if vin_normalizado in unidades
                else None
            )
            if vin_normalizado and unidad_id is None:
                append_error(fila, "INVENTARIO_UNIDAD_NO_CONCILIADA")
            importe_original = cell(workbook, hoja, fila.source.row, 4)
            costo_original = cell(workbook, hoja, fila.source.row, 6)
            importe = parse_amount(importe_original)
            costo = parse_amount(costo_original)
            if importe_original not in (None, "") and importe is None:
                append_error(fila, "INVENTARIO_IMPORTE_INVALIDO")
            if costo_original not in (None, "") and costo is None:
                append_error(fila, "INVENTARIO_COSTO_INVALIDO")
            registro_id = deterministic_id(
                "registro-inventario-historico",
                f"{organizacion_id}:{fila_importacion_id}",
            )
            registros.append(
                (
                    registro_id,
                    organizacion_id,
                    sucursal_id,
                    fila_importacion_id,
                    hoja,
                    fecha,
                    safe_text(cell(workbook, hoja, fila.source.row, 2)),
                    vin_mostrado,
                    vin_normalizado,
                    unidad_id,
                    limited_text(cell(workbook, hoja, fila.source.row, 5), 120),
                    importe,
                    costo,
                    Jsonb(
                        inferred_data(
                            origen="excel_luma",
                            fecha_no_informada=fecha is None,
                            columna_7_original=cell(
                                workbook, hoja, fila.source.row, 7
                            ),
                            columna_8_original=cell(
                                workbook, hoja, fila.source.row, 8
                            ),
                        )
                    ),
                )
            )
            filas.append(fila)

    registros_reales: dict[uuid.UUID, uuid.UUID] = {}
    with connection.cursor() as cursor:
        ejecutar_lote(
            cursor,
            """
            INSERT INTO registros_inventario_importados (
              id,
              organizacion_id,
              sucursal_id,
              fila_importacion_id,
              hoja_origen,
              fecha_registro,
              descripcion,
              vin_mostrado,
              vin_normalizado,
              unidad_vehiculo_id,
              estado_original,
              importe,
              costo,
              datos_inferidos
            )
            VALUES (
              %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
            )
            ON CONFLICT (organizacion_id, fila_importacion_id) DO UPDATE SET
              sucursal_id = EXCLUDED.sucursal_id,
              hoja_origen = EXCLUDED.hoja_origen,
              fecha_registro = EXCLUDED.fecha_registro,
              descripcion = EXCLUDED.descripcion,
              vin_mostrado = EXCLUDED.vin_mostrado,
              vin_normalizado = EXCLUDED.vin_normalizado,
              unidad_vehiculo_id = EXCLUDED.unidad_vehiculo_id,
              estado_original = EXCLUDED.estado_original,
              importe = EXCLUDED.importe,
              costo = EXCLUDED.costo,
              datos_inferidos = EXCLUDED.datos_inferidos,
              actualizado_en = now()
            """,
            registros,
        )
        if filas:
            cursor.execute(
                """
                SELECT id, fila_importacion_id
                FROM registros_inventario_importados
                WHERE organizacion_id = %s
                  AND fila_importacion_id = ANY(%s)
                """,
                (organizacion_id, [fila.fila_importacion_id for fila in filas]),
            )
            registros_reales = {
                fila_id: registro_id
                for registro_id, fila_id in cursor.fetchall()
            }
    for fila in filas:
        registro_id = registros_reales.get(fila.fila_importacion_id)
        if registro_id is None:
            append_error(fila, "INVENTARIO_REGISTRO_NO_PERSISTIDO")
        else:
            mark_target(
                staged_by_source,
                fila.source,
                "registros_inventario_importados",
                registro_id,
            )
    return {"registros_inventario_importados": len(registros_reales)}


def importar_datos_prueba(
    connection: psycopg.Connection[Any],
    workbook: openpyxl.Workbook,
    staged: list[StagedRow],
    staged_by_source: dict[SourceRef, StagedRow],
    organizacion_id: uuid.UUID,
    sucursal_predeterminada: str,
) -> dict[str, int]:
    sucursal_id = resolver_sucursal(
        connection, organizacion_id, sucursal_predeterminada
    )
    actor_sistema_id = asegurar_actor_sistema_importado(
        connection, organizacion_id, sucursal_id
    )
    personal_por_nombre = cargar_personal_por_nombre(connection, organizacion_id)
    cuentas_requeridas = cuentas_historicas_requeridas(workbook, staged)
    cuentas_historicas = asegurar_cuentas_historicas(
        connection,
        organizacion_id,
        sucursal_id,
        cuentas_requeridas,
        personal_por_nombre,
    )
    resumen: dict[str, int] = {
        "cuentas_historicas": len(cuentas_historicas),
        "clientes_canonicos_operaciones": importar_clientes_canonicos_para_operaciones(
            connection,
            workbook,
            staged,
            staged_by_source,
            organizacion_id,
        ),
    }
    for resultado in (
        importar_operaciones_historicas(
            connection,
            workbook,
            staged,
            staged_by_source,
            organizacion_id,
            sucursal_id,
            actor_sistema_id,
        ),
        importar_ingresos_historicos(
            connection,
            workbook,
            staged,
            staged_by_source,
            organizacion_id,
            sucursal_id,
            actor_sistema_id,
            cuentas_historicas,
        ),
        importar_gastos_ventas_historicos(
            connection,
            workbook,
            staged,
            staged_by_source,
            organizacion_id,
            sucursal_id,
            actor_sistema_id,
            cuentas_historicas,
        ),
        importar_pagos_formulario_historicos(
            connection,
            workbook,
            staged,
            staged_by_source,
            organizacion_id,
            sucursal_id,
            actor_sistema_id,
            cuentas_historicas,
        ),
        importar_polizas_historicas(
            connection,
            workbook,
            staged,
            staged_by_source,
            organizacion_id,
            sucursal_id,
        ),
        importar_prospectos_historicos(
            connection,
            workbook,
            staged,
            staged_by_source,
            organizacion_id,
            sucursal_id,
        ),
        importar_registros_inventario_historicos(
            connection,
            workbook,
            staged,
            staged_by_source,
            organizacion_id,
            sucursal_id,
        ),
    ):
        resumen.update(resultado)
    return resumen


def classify_rows(
    staged: list[StagedRow], datos_prueba: bool = False
) -> None:
    for row in staged:
        if row.estado == "OMITIDA":
            continue
        if not datos_prueba:
            deferred_error = DEFERRED_SHEETS.get(row.source.sheet)
            if deferred_error and deferred_error not in row.errors:
                row.errors.append(deferred_error)
            if row.source.sheet in {"Clientes.", "VENTAS"}:
                append_error(row, "OPERACION_CONCILIACION_REQUERIDA")
        elif row.source.sheet == "lucho":
            append_error(row, DEFERRED_SHEETS["lucho"])
        if row.errors:
            row.estado = "EN_CUARENTENA"
        elif row.targets:
            row.estado = "IMPORTADA"
        else:
            row.estado = "NORMALIZADA"


def finalize_batch(
    connection: psycopg.Connection[Any],
    lote_id: uuid.UUID,
    staged: list[StagedRow],
    datos_prueba: bool = False,
) -> None:
    classify_rows(staged, datos_prueba)
    imported = sum(row.estado == "IMPORTADA" for row in staged)
    quarantined = sum(row.estado == "EN_CUARENTENA" for row in staged)
    with connection.cursor() as cursor:
        ejecutar_lote(
            cursor,
            """
            UPDATE filas_importacion
            SET estado = %s,
                codigos_error = %s,
                referencias_destino = %s,
                carga_normalizada = %s,
                actualizado_en = now()
            WHERE lote_id = %s
              AND nombre_hoja = %s
              AND nombre_bloque = %s
              AND fila_origen = %s
            """,
            [
                (
                    row.estado,
                    Jsonb(sorted(set(row.errors))),
                    Jsonb(row.targets),
                    (
                        Jsonb(
                            {
                                "normalizada": True,
                                "modo": (
                                    "datos_prueba"
                                    if datos_prueba
                                    else "maestros_seguros"
                                ),
                            }
                        )
                        if row.targets
                        else None
                    ),
                    lote_id,
                    row.source.sheet,
                    row.source.block,
                    row.source.row,
                )
                for row in staged
            ],
        )
        cursor.execute(
            """
            UPDATE lotes_importacion
            SET estado = 'COMPLETADO',
                filas_importadas = %s,
                filas_cuarentena = %s,
                finalizado_en = now(),
                actualizado_en = now()
            WHERE id = %s
            """,
            (imported, quarantined, lote_id),
        )


def summary(
    staged: list[StagedRow],
    clientes: list[CustomerCandidate],
    staff: list[StaffCandidate],
    proveedores: list[SupplierCandidate],
    vehicles: list[VehicleCandidate],
    apply: bool,
    datos_prueba: bool = False,
    resultados_datos_prueba: dict[str, int] | None = None,
) -> dict[str, Any]:
    sheets = Counter(row.source.sheet for row in staged)
    errors = Counter(error for row in staged for error in row.errors)
    resultado: dict[str, Any] = {
        "modo": "aplicar" if apply else "simulacion",
        "datos_prueba": datos_prueba,
        "filas_staging": len(staged),
        "filas_por_hoja": dict(sorted(sheets.items())),
        "candidatos": {
            "clientes": len(clientes),
            "personal": len(staff),
            "proveedores": len(proveedores),
            "filas_catalogo": len(vehicles),
            "unidades_con_vin_valido": sum(
                vehicle.vin_normalizado is not None for vehicle in vehicles
            ),
        },
        "incidencias": dict(sorted(errors.items())),
    }
    if resultados_datos_prueba is not None:
        resultado["datos_prueba_importados"] = dict(
            sorted(resultados_datos_prueba.items())
        )
    return resultado


def main() -> int:
    args = parse_args()
    workbook = args.libro.resolve()
    if not workbook.is_file():
        print(f"Libro no encontrado: {workbook}", file=sys.stderr)
        return 2
    if workbook.suffix.lower() != ".xlsx":
        print("Solo se admiten libros .xlsx.", file=sys.stderr)
        return 2
    if args.solo_staging and not args.aplicar:
        print("--solo-staging requiere --aplicar.", file=sys.stderr)
        return 2
    if args.datos_prueba:
        if not args.aplicar:
            print("--datos-prueba requiere --aplicar.", file=sys.stderr)
            return 2
        if args.solo_staging:
            print("--datos-prueba no es compatible con --solo-staging.", file=sys.stderr)
            return 2
        if not safe_text(args.organizacion):
            print(
                "--datos-prueba requiere --organizacion explicita.",
                file=sys.stderr,
            )
            return 2
        if not safe_text(args.sucursal_predeterminada):
            print(
                "--datos-prueba requiere --sucursal-predeterminada explicita.",
                file=sys.stderr,
            )
            return 2
        if normalize_text(args.organizacion) != "luma_central":
            print(
                "--datos-prueba solo admite --organizacion LUMA_CENTRAL.",
                file=sys.stderr,
            )
            return 2
        if normalize_text(args.sucursal_predeterminada) != "san miguel":
            print(
                "--datos-prueba solo admite la sucursal San Miguel.",
                file=sys.stderr,
            )
            return 2
        print(
            "[DatosPrueba] Se importaran datos reales solo como prueba en "
            "LUMA_CENTRAL / San Miguel.",
            file=sys.stderr,
        )
    codigo_organizacion_solicitada = args.organizacion or "LUMA_CENTRAL"

    workbook_hash = sha256_file(workbook)
    warnings.filterwarnings(
        "ignore",
        message=r".*pivotCacheDefinition.*invalid dependency definitions.*",
        category=UserWarning,
        module=r"openpyxl\.packaging\.relationship",
    )
    workbook_values = openpyxl.load_workbook(
        workbook,
        read_only=False,
        data_only=True,
    )
    workbook_formulas = openpyxl.load_workbook(
        workbook,
        read_only=False,
        data_only=False,
    )
    staged = load_staged_rows(workbook_values, workbook_formulas)
    print(
        f"[Preparacion] {len(staged)} filas fuente listas para staging.",
        file=sys.stderr,
    )
    staged_by_source = {row.source: row for row in staged}
    clientes = extract_customers(workbook_values, staged)
    flag_customer_identity_conflicts(clientes, staged_by_source)
    staff = extract_staff(workbook_values, staged)
    proveedores = extract_suppliers(workbook_values, staged)
    vehicles = extract_vehicles(workbook_values, staged)

    if not args.aplicar:
        classify_rows(staged, args.datos_prueba)
        print(
            json.dumps(
                summary(
                    staged,
                    clientes,
                    staff,
                    proveedores,
                    vehicles,
                    False,
                    args.datos_prueba,
                ),
                indent=2,
                sort_keys=True,
            )
        )
        return 0

    database_url = os.environ.get(args.variable_entorno_base_datos)
    if not database_url:
        print(
            f"La variable de entorno {args.variable_entorno_base_datos} no esta configurada.",
            file=sys.stderr,
        )
        return 2

    # autocommit permite que el bloque siguiente sea la transaccion externa;
    # las transacciones internas del importador pasan a ser puntos de guardado.
    with psycopg.connect(database_url, autocommit=True) as connection:
        ensure_import_schema(connection, args.datos_prueba)
        with connection.transaction():
            organizacion_id, codigo_organizacion = resolve_organizacion(
                connection, codigo_organizacion_solicitada
            )
            sucursal_datos_prueba_id = (
                resolver_sucursal(
                    connection,
                    organizacion_id,
                    args.sucursal_predeterminada,
                )
                if args.datos_prueba
                else None
            )
            print("[Staging] Insertando o actualizando filas de origen.", file=sys.stderr)
            lote_id = stage_rows(
                connection,
                workbook,
                workbook_hash,
                staged,
                organizacion_id,
            )
            resultados_datos_prueba: dict[str, int] | None = None
            if not args.solo_staging:
                cantidad_personal = import_staff(
                    connection,
                    staff,
                    staged_by_source,
                    organizacion_id,
                    codigo_organizacion,
                    sucursal_datos_prueba_id,
                )
                print(
                    f"[Maestros] {cantidad_personal} filas de personal procesadas.",
                    file=sys.stderr,
                )
                supplier_ids = import_suppliers(
                    connection,
                    proveedores,
                    staged_by_source,
                    organizacion_id,
                    codigo_organizacion,
                )
                cantidad_clientes = import_customers(
                    connection,
                    clientes,
                    staged_by_source,
                    organizacion_id,
                    codigo_organizacion,
                )
                catalogos, unidades, unidades_cuarentena = import_catalog_and_units(
                    connection,
                    vehicles,
                    staged_by_source,
                    supplier_ids,
                    args.sucursal_predeterminada,
                    organizacion_id,
                    codigo_organizacion,
                    args.datos_prueba,
                )
                print(
                    "[Maestros] "
                    f"{cantidad_clientes} clientes, {catalogos} catalogos y "
                    f"{unidades} unidades procesadas "
                    f"({unidades_cuarentena} en cuarentena).",
                    file=sys.stderr,
                )
                if args.datos_prueba:
                    print(
                        "[DatosPrueba] Importando operaciones, finanzas e historicos.",
                        file=sys.stderr,
                    )
                    resultados_datos_prueba = importar_datos_prueba(
                        connection,
                        workbook_values,
                        staged,
                        staged_by_source,
                        organizacion_id,
                        args.sucursal_predeterminada,
                    )
                    print(
                        "[DatosPrueba] "
                        + ", ".join(
                            f"{clave}={valor}"
                            for clave, valor in sorted(
                                resultados_datos_prueba.items()
                            )
                        ),
                        file=sys.stderr,
                    )
            print("[Finalizacion] Actualizando estados y referencias de origen.", file=sys.stderr)
            finalize_batch(connection, lote_id, staged, args.datos_prueba)

    print(
        json.dumps(
            summary(
                staged,
                clientes,
                staff,
                proveedores,
                vehicles,
                True,
                args.datos_prueba,
                resultados_datos_prueba,
            ),
            indent=2,
            sort_keys=True,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
