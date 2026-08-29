"""Pruebas locales sin base de datos para reglas de importacion historica."""

from __future__ import annotations

import unittest
from pathlib import Path
from unittest.mock import patch

import openpyxl

from importar_excel_luma import (
    IMPORT_DATABASE_ENV_VAR,
    SourceRef,
    StagedRow,
    clasificar_venta_por_documento,
    es_transferencia_origen,
    parse_amount,
    parse_args,
    positive_amount,
)

REPOSITORY_ROOT = Path(__file__).resolve().parents[1]
LEGACY_MIGRATION = (
    REPOSITORY_ROOT
    / "prisma"
    / "migrations"
    / "20260829130000_legacy_import_support"
    / "migration.sql"
)
PRISMA_SCHEMA = REPOSITORY_ROOT / "prisma" / "schema.prisma"
CATALOG_TRIGGER_MIGRATION = (
    REPOSITORY_ROOT
    / "prisma"
    / "migrations"
    / "20260829131000_protect_shared_catalog_inserts"
    / "migration.sql"
)


class ReglasImportacionHistoricaTests(unittest.TestCase):
    def test_importador_usa_direct_url_por_defecto(self) -> None:
        with patch(
            "sys.argv",
            ["importar_excel_luma.py", "--libro", "datos.xlsx"],
        ):
            argumentos = parse_args()

        self.assertEqual(
            argumentos.variable_entorno_base_datos,
            IMPORT_DATABASE_ENV_VAR,
        )
        self.assertEqual(IMPORT_DATABASE_ENV_VAR, "DIRECT_URL")

    def test_importes_argentinos_y_positivos(self) -> None:
        self.assertEqual(str(parse_amount("$ 1.234,50")), "1234.50")
        self.assertEqual(str(parse_amount("1,234.50")), "1234.50")
        self.assertIsNone(positive_amount(0))
        self.assertIsNone(parse_amount("-1"))

    def test_transferencias_no_se_clasifican_como_ingresos_de_caja(self) -> None:
        self.assertTrue(es_transferencia_origen("Transferencia bancaria"))
        self.assertFalse(es_transferencia_origen("Credito"))

    def test_documento_duplicado_se_resuelve_solo_por_vin(self) -> None:
        libro = openpyxl.Workbook()
        clientes = libro.active
        clientes.title = "Clientes."
        ventas = libro.create_sheet("VENTAS")
        clientes.cell(4, 2).value = "12345678"
        clientes.cell(4, 15).value = "PRUEBA-UNO-123"
        clientes.cell(5, 2).value = "12345678"
        clientes.cell(5, 15).value = "PRUEBA-DOS-456"
        ventas.cell(4, 2).value = "PRUEBA-DOS-456"
        ventas.cell(4, 6).value = "12345678"

        primera = StagedRow(SourceRef("Clientes.", 4), {}, "a")
        segunda = StagedRow(SourceRef("Clientes.", 5), {}, "b")
        venta = StagedRow(SourceRef("VENTAS", 4), {}, "c")

        resultado = clasificar_venta_por_documento(
            libro, venta, {"12345678": [primera, segunda]}
        )

        self.assertIs(resultado, segunda)
        self.assertEqual(venta.errors, [])

    def test_documento_duplicado_sin_vin_no_se_infiere(self) -> None:
        libro = openpyxl.Workbook()
        clientes = libro.active
        clientes.title = "Clientes."
        ventas = libro.create_sheet("VENTAS")
        clientes.cell(4, 2).value = "12345678"
        clientes.cell(4, 15).value = "PRUEBA-UNO-123"
        clientes.cell(5, 2).value = "12345678"
        clientes.cell(5, 15).value = "PRUEBA-DOS-456"
        ventas.cell(4, 2).value = "PRUEBA-TRES-789"
        ventas.cell(4, 6).value = "12345678"

        primera = StagedRow(SourceRef("Clientes.", 4), {}, "a")
        segunda = StagedRow(SourceRef("Clientes.", 5), {}, "b")
        venta = StagedRow(SourceRef("VENTAS", 4), {}, "c")

        resultado = clasificar_venta_por_documento(
            libro, venta, {"12345678": [primera, segunda]}
        )

        self.assertIsNone(resultado)
        self.assertIn("VENTA_CLIENTE_DOCUMENTO_AMBIGUO", venta.errors)

    def test_migracion_prisma_porta_superficies_legacy(self) -> None:
        sql = LEGACY_MIGRATION.read_text(encoding="utf-8")

        for table in (
            "ingresos",
            "polizas_seguros",
            "prospectos",
            "registros_inventario_importados",
        ):
            self.assertIn(f'CREATE TABLE IF NOT EXISTS "{table}"', sql)
            self.assertIn(f'ALTER TABLE "{table}" FORCE ROW LEVEL SECURITY', sql)

        self.assertIn("(transferencia_id IS NOT NULL)::integer", sql)
        self.assertNotIn("transferenciaencia_id", sql)
        trigger_sql = CATALOG_TRIGGER_MIGRATION.read_text(encoding="utf-8")
        for table in ("marcas_vehiculos", "modelos_vehiculos"):
            self.assertIn(
                f'BEFORE INSERT OR UPDATE OR DELETE ON "{table}"',
                trigger_sql,
            )
        self.assertIn(
            'EXECUTE FUNCTION "luma_proteger_catalogo_compartido"()',
            trigger_sql,
        )

    def test_schema_conserva_nombre_historico_de_constraint(self) -> None:
        schema = PRISMA_SCHEMA.read_text(encoding="utf-8")

        self.assertIn(
            'map: "movimientos_caja_transferenciaencia_id_fkey"',
            schema,
        )


if __name__ == "__main__":
    unittest.main()
